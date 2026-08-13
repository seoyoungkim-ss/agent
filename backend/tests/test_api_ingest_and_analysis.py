import datetime as dt
import io

import pytest

from app.config import get_settings

AUTH_HEADERS = {"Authorization": f"Bearer {get_settings().ingest_api_token}"}

MONDAY = dt.date(2026, 7, 20)


def _ingest_weekly_menu(client):
    rows = [
        {
            "plan_date": MONDAY.isoformat(),
            "meal_type": "중식",
            "corner_name": "한식",
            "menu_name": "제육볶음",
            "menu_role": "메인",
            "source_row_raw": "제육볶음\n계란후라이",
        },
        {
            "plan_date": MONDAY.isoformat(),
            "meal_type": "중식",
            "corner_name": "한식",
            "menu_name": "계란후라이",
            "menu_role": "부찬",
            "source_row_raw": "제육볶음\n계란후라이",
        },
    ]
    resp = client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text
    return resp.json()


def _ingest_meal_log(
    client,
    employee_id: str,
    taste: str,
    comment: str | None = None,
    company_name: str | None = None,
    eaten_date: dt.date = MONDAY,
    menu_name: str | None = None,
    corner_name: str = "한식",
):
    rows = [
        {
            "eaten_at": dt.datetime.combine(eaten_date, dt.time(11, 52, 0)).isoformat(),
            "employee_id": employee_id,
            "meal_type": "중식",
            "corner_name": corner_name,
            "taste_score": taste,
            "comment": comment,
            "company_name": company_name,
            "menu_name": menu_name,
        }
    ]
    resp = client.post("/api/ingest/meal-log", json={"rows": rows}, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text
    return resp.json()


def test_ingest_requires_token(client):
    resp = client.post("/api/ingest/weekly-menu", json={"rows": []})
    assert resp.status_code == 401


def test_ingest_weekly_menu_creates_menus_and_corners(client):
    result = _ingest_weekly_menu(client)
    assert result["inserted"] == 2
    assert result["new_menus"] == 2
    assert result["new_corners"] == 1


def test_ingest_meal_log_links_to_main_menu_snapshot(client, db_session):
    _ingest_weekly_menu(client)
    _ingest_meal_log(client, "E12345", "맛남", "맛있어요")

    from app.models.logs import MealLog

    log = db_session.query(MealLog).filter_by(employee_id="E12345").one()
    assert log.menu_id is not None  # 코너의 메인 메뉴(제육볶음)와 연결돼야 함

    from app.models.master import MenuMaster

    menu = db_session.query(MenuMaster).filter_by(menu_id=log.menu_id).one()
    assert menu.menu_name == "제육볶음"


def test_menu_food_vector_auto_tagged_by_rule_on_ingest(client, db_session):
    _ingest_weekly_menu(client)

    from app.models.master import MenuMaster

    menu = db_session.query(MenuMaster).filter_by(menu_name="제육볶음").one()
    assert menu.food_vector is not None  # "제육"(protein) + "볶음"(oily) 키워드에 걸림
    assert menu.food_vector_source.value == "규칙기반"


def test_menu_food_vector_stays_untagged_when_no_rule_matches(client, db_session):
    _ingest_meal_log(client, "E55555", "맛남", menu_name="모듬과일")

    from app.models.master import MenuMaster

    menu = db_session.query(MenuMaster).filter_by(menu_name="모듬과일").one()
    assert menu.food_vector is None
    assert menu.food_vector_source is None


def test_list_menu_food_vectors_endpoint(client):
    _ingest_weekly_menu(client)
    _ingest_meal_log(client, "E11111", "맛남")  # 코너의 메인 메뉴(제육볶음)와 자동 연결, 코너 "한식"
    _ingest_meal_log(client, "E55555", "맛남", menu_name="모듬과일", corner_name="분식")

    resp = client.get("/api/analysis/menus/food-vectors")
    assert resp.status_code == 200
    rows = resp.json()
    names = {row["menu_name"] for row in rows}
    assert "제육볶음" in names
    assert "모듬과일" in names
    assert "계란후라이" in names
    jeyuk = next(r for r in rows if r["menu_name"] == "제육볶음")
    assert jeyuk["corner_name"] == "한식"  # meal_log에서 실제 취식된 코너 기준
    moduem = next(r for r in rows if r["menu_name"] == "모듬과일")
    assert moduem["corner_name"] == "분식"  # 실제로 이 코너에서 취식됨
    gyeranhurai = next(r for r in rows if r["menu_name"] == "계란후라이")
    assert gyeranhurai["corner_name"] is None  # meal_log에 취식 기록이 없는 메뉴(부찬) — 코너 미배정

    resp_untagged = client.get("/api/analysis/menus/food-vectors", params={"untagged_only": True})
    untagged_names = {row["menu_name"] for row in resp_untagged.json()}
    assert untagged_names == {"모듬과일"}


def test_update_menu_food_vector_manual_override(client, db_session):
    _ingest_meal_log(client, "E55555", "맛남", menu_name="모듬과일")
    from app.models.master import MenuMaster

    menu = db_session.query(MenuMaster).filter_by(menu_name="모듬과일").one()

    vector = [0.1] * 10
    resp = client.put(f"/api/analysis/menus/{menu.menu_id}/food-vector", json={"vector": vector})
    assert resp.status_code == 200
    assert resp.json()["source"] == "관리자수동"

    db_session.refresh(menu)
    assert menu.food_vector_source.value == "관리자수동"
    assert list(menu.food_vector) == vector

    resp_bad_length = client.put(
        f"/api/analysis/menus/{menu.menu_id}/food-vector", json={"vector": [0.1] * 3}
    )
    assert resp_bad_length.status_code == 400

    resp_out_of_range = client.put(
        f"/api/analysis/menus/{menu.menu_id}/food-vector", json={"vector": [1.5] * 10}
    )
    assert resp_out_of_range.status_code == 400

    resp_missing = client.put("/api/analysis/menus/999999/food-vector", json={"vector": [0.1] * 10})
    assert resp_missing.status_code == 404


def test_tag_menus_with_llm_leaves_untagged_when_llm_unconfigured(client):
    _ingest_meal_log(client, "E55555", "맛남", menu_name="모듬과일")

    resp = client.post("/api/analysis/menus/tag-with-llm")
    assert resp.status_code == 200
    # 사내 LLM 미설정 환경에서는 모의 응답이 벡터 형식으로 파싱되지 않으므로 0건이어야 함
    assert resp.json()["tagged_menus"] == 0


def test_health_endpoint(client):
    resp = client.get("/health")
    assert resp.status_code == 200
    assert resp.json() == {"status": "ok"}


def test_menu_performance_recompute_and_read(client):
    _ingest_weekly_menu(client)
    _ingest_meal_log(client, "E1", "맛남")
    _ingest_meal_log(client, "E2", "맛남")
    _ingest_meal_log(client, "E3", "개선")

    resp = client.post(
        "/api/analysis/menu-performance/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    assert resp.json()["updated_menus"] >= 1

    resp = client.get(
        "/api/analysis/menu-performance",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    rows = resp.json()
    jeyuk = next(r for r in rows if r["menu_name"] == "제육볶음")
    assert jeyuk["evaluation_count"] == 3
    assert jeyuk["quadrant"] is not None
    assert jeyuk["corner_name"] == "한식"  # meal_log에서 실제 취식된 코너 기준


def test_menu_performance_recompute_excludes_take_out_placeholder_menus(client):
    _ingest_weekly_menu(client)
    _ingest_meal_log(client, "E1", "맛남")
    _ingest_meal_log(client, "E2", "맛남")
    _ingest_meal_log(client, "E3", "개선")
    # 테이크아웃 플레이스홀더 메뉴 — 4분면 집계 자체에서 빠져야 함(중앙값도 왜곡 안 함)
    _ingest_meal_log(client, "E4", "맛남", menu_name="선택형 Take out", corner_name="Take Out")
    _ingest_meal_log(client, "E5", "맛남", menu_name="(포장)메디쏠라", corner_name="Take Out")

    resp = client.post(
        "/api/analysis/menu-performance/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200

    resp = client.get(
        "/api/analysis/menu-performance",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    menu_names = {r["menu_name"] for r in resp.json()}
    assert "선택형 Take out" not in menu_names
    assert "(포장)메디쏠라" not in menu_names


def test_menu_performance_loyal_following_overrides_low_demand_to_hidden_gem(client):
    # "로열티메뉴"는 3번만 나오고 총 식수도 10명뿐이라 수요는 낮지만(대조군
    # "인기메뉴테스트" 대비), 그 메뉴가 나올 때마다(3번 다) 챙겨 먹는 고정
    # 고객 2명이 있으면 만족도가 낮아도(전부 "개선") 퇴출후보 대신 숨은강자로
    # 분류돼야 한다(2026-07 사용자 요청).
    loyalty_dates = [MONDAY, MONDAY + dt.timedelta(days=7), MONDAY + dt.timedelta(days=14)]
    for d in loyalty_dates:
        _ingest_meal_log(client, "EL1", "개선", eaten_date=d, menu_name="로열티메뉴", corner_name="한식")
        _ingest_meal_log(client, "EL2", "개선", eaten_date=d, menu_name="로열티메뉴", corner_name="한식")
    # 한 번씩만 찾는 사람들 — 로열티 조건(최소 주문횟수 2회) 미달
    _ingest_meal_log(client, "EC1", "개선", eaten_date=loyalty_dates[0], menu_name="로열티메뉴", corner_name="한식")
    _ingest_meal_log(client, "EC2", "개선", eaten_date=loyalty_dates[0], menu_name="로열티메뉴", corner_name="한식")
    _ingest_meal_log(client, "EC3", "개선", eaten_date=loyalty_dates[1], menu_name="로열티메뉴", corner_name="한식")
    _ingest_meal_log(client, "EC4", "개선", eaten_date=loyalty_dates[2], menu_name="로열티메뉴", corner_name="한식")

    # 대조군 — 수요를 확실히 높게 만들어 로열티메뉴가 저수요 분면에 들어가게 함
    for i in range(20):
        _ingest_meal_log(client, f"EP{i}", "맛남", eaten_date=MONDAY, menu_name="인기메뉴테스트", corner_name="일품")

    period_end = MONDAY + dt.timedelta(days=14)
    resp = client.post(
        "/api/analysis/menu-performance/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": period_end.isoformat()},
    )
    assert resp.status_code == 200

    resp = client.get(
        "/api/analysis/menu-performance",
        params={"period_start": MONDAY.isoformat(), "period_end": period_end.isoformat()},
    )
    row = next(r for r in resp.json() if r["menu_name"] == "로열티메뉴")
    assert row["has_loyal_following"] is True
    assert row["quadrant"] == "숨은강자"


def test_menu_performance_satisfaction_trend_detects_recent_decline(client):
    # 직전(prior) 30일엔 만족도가 높았다가 최근(recent) 30일엔 뚝 떨어진 메뉴 —
    # satisfaction_trend가 "하락"으로 잡혀야 한다(2026-07 사용자 요청,
    # menu_trend_window_days=30 기본값 기준 period_end 앵커).
    period_end = dt.date(2026, 7, 31)
    prior_date = period_end - dt.timedelta(days=45)  # 직전 30일 구간 안
    recent_date = period_end - dt.timedelta(days=10)  # 최근 30일 구간 안
    period_start = period_end - dt.timedelta(days=89)

    for i in range(10):
        _ingest_meal_log(client, f"EP{i}", "맛남", eaten_date=prior_date, menu_name="하락메뉴", corner_name="한식")
    for i in range(10):
        _ingest_meal_log(client, f"ER{i}", "개선", eaten_date=recent_date, menu_name="하락메뉴", corner_name="한식")

    resp = client.post(
        "/api/analysis/menu-performance/recompute",
        params={"period_start": period_start.isoformat(), "period_end": period_end.isoformat()},
    )
    assert resp.status_code == 200

    resp = client.get(
        "/api/analysis/menu-performance",
        params={"period_start": period_start.isoformat(), "period_end": period_end.isoformat()},
    )
    row = next(r for r in resp.json() if r["menu_name"] == "하락메뉴")
    assert row["satisfaction_trend"] == "하락"


def test_menu_performance_by_meal_type_includes_trend_and_loyalty_fields(client):
    _ingest_meal_log_with_meal_type(client, "E1", "맛남", "조식", "토스트")
    _ingest_meal_log_with_meal_type(client, "E2", "맛남", "조식", "토스트")

    resp = client.get(
        "/api/analysis/menu-performance/by-meal-type",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "meal_type": "조식"},
    )
    assert resp.status_code == 200
    row = next(r for r in resp.json() if r["menu_name"] == "토스트")
    assert row["satisfaction_trend"] in ("상승", "유지", "하락")
    assert isinstance(row["has_loyal_following"], bool)


def test_menu_highlights_detects_rising_menu_and_new_menu_reaction(client):
    _ingest_weekly_menu(client)  # 제육볶음/계란후라이, 한식, MONDAY

    prior_week = MONDAY - dt.timedelta(days=14)
    for i in range(3):
        _ingest_meal_log(client, f"P{i}", "개선", eaten_date=prior_week, menu_name="제육볶음", corner_name="한식")
    for i in range(3):
        _ingest_meal_log(client, f"R{i}", "맛남", eaten_date=MONDAY, menu_name="제육볶음", corner_name="한식")

    # 신메뉴 — 처음 등장하는 메뉴라 자동으로 is_new_menu=True로 찍힘
    resp = client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                {
                    "plan_date": MONDAY.isoformat(),
                    "meal_type": "중식",
                    "corner_name": "한식",
                    "menu_name": "신메뉴테스트",
                    "menu_role": "메인",
                    "source_row_raw": "신메뉴테스트",
                }
            ]
        },
        headers=AUTH_HEADERS,
    )
    assert resp.status_code == 200
    _ingest_meal_log(client, "N1", "맛남", eaten_date=MONDAY, menu_name="신메뉴테스트", corner_name="한식")

    resp = client.get("/api/dashboard/menu-highlights")
    assert resp.status_code == 200
    body = resp.json()

    rising_names = {r["menu_name"] for r in body["rising"]}
    falling_names = {r["menu_name"] for r in body["falling"]}
    assert "제육볶음" in rising_names
    assert "제육볶음" not in falling_names
    # date는 마지막으로 나온 주(MONDAY)의 월요일 — 홈 하이라이트 카드 날짜 표시용(2026-08)
    rising_entry = next(r for r in body["rising"] if r["menu_name"] == "제육볶음")
    assert rising_entry["date"] == MONDAY.isoformat()

    new_menu_names = {r["menu_name"] for r in body["new_menus"]}
    assert "신메뉴테스트" in new_menu_names
    new_menu_entry = next(r for r in body["new_menus"] if r["menu_name"] == "신메뉴테스트")
    assert new_menu_entry["evaluation_count"] == 1
    assert new_menu_entry["days_since_introduction"] == (dt.date.today() - MONDAY).days
    assert new_menu_entry["needs_attention"] is False  # 이미 평가가 있음


def test_menu_highlights_flags_new_menu_needing_attention_when_unevaluated(client):
    # 도입 후 오래됐는데(NEW_MENU_WINDOW_DAYS=30 이내지만 7일은 넘김) 평가가
    # 하나도 없으면 관심 유도가 필요하다는 신호(needs_attention)를 켠다.
    stale_date = dt.date.today() - dt.timedelta(days=10)
    resp = client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                {
                    "plan_date": stale_date.isoformat(),
                    "meal_type": "중식",
                    "corner_name": "한식",
                    "menu_name": "무관심신메뉴",
                    "menu_role": "메인",
                    "source_row_raw": "무관심신메뉴",
                }
            ]
        },
        headers=AUTH_HEADERS,
    )
    assert resp.status_code == 200

    resp = client.get("/api/dashboard/menu-highlights")
    assert resp.status_code == 200
    entry = next(r for r in resp.json()["new_menus"] if r["menu_name"] == "무관심신메뉴")
    assert entry["evaluation_count"] == 0
    assert entry["days_since_introduction"] == 10
    assert entry["needs_attention"] is True


def test_menu_highlights_excludes_side_dish_from_new_menus(client):
    # 신메뉴는 메인메뉴만 의미가 있다 — 같은 날 부찬으로 처음 등장한 메뉴는
    # is_new_menu=True로 찍혀도 하이라이트에 안 떠야 한다.
    resp = client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                {
                    "plan_date": MONDAY.isoformat(),
                    "meal_type": "중식",
                    "corner_name": "한식",
                    "menu_name": "신메뉴부찬",
                    "menu_role": "부찬",
                    "source_row_raw": "신메뉴부찬",
                }
            ]
        },
        headers=AUTH_HEADERS,
    )
    assert resp.status_code == 200

    resp = client.get("/api/dashboard/menu-highlights")
    assert resp.status_code == 200
    assert "신메뉴부찬" not in {r["menu_name"] for r in resp.json()["new_menus"]}

    # 관리자가 이 부찬을 수동으로 신메뉴 지정해도 하이라이트엔 안 떠야 한다.
    resp = client.put("/api/analysis/menus/new-menu-status", json={"menu_name": "신메뉴부찬", "is_new": True})
    assert resp.status_code == 200, resp.text
    resp = client.get("/api/dashboard/menu-highlights")
    assert "신메뉴부찬" not in {r["menu_name"] for r in resp.json()["new_menus"]}


def test_new_menu_status_manual_add_bypasses_auto_window(client):
    # meal-log만으로 생긴 메뉴는 weekly_menu_plan.is_new_menu가 전혀 안 찍히므로
    # 자동판정으로는 "신메뉴 반응"에 절대 안 뜬다 — 관리자가 직접 등록하면 떠야 함.
    _ingest_meal_log(client, "E1", "맛남", menu_name="관리자등록메뉴", corner_name="한식")

    resp = client.get("/api/dashboard/menu-highlights")
    assert "관리자등록메뉴" not in {r["menu_name"] for r in resp.json()["new_menus"]}

    resp = client.put(
        "/api/analysis/menus/new-menu-status", json={"menu_name": "관리자등록메뉴", "is_new": True}
    )
    assert resp.status_code == 200, resp.text

    resp = client.get("/api/dashboard/menu-highlights")
    entry = next(r for r in resp.json()["new_menus"] if r["menu_name"] == "관리자등록메뉴")
    assert entry["is_manual"] is True
    assert entry["days_since_introduction"] == 0
    assert entry["corner_name"] == "한식"


def test_new_menu_status_manual_remove_hides_auto_detected_menu(client):
    resp = client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                {
                    "plan_date": MONDAY.isoformat(),
                    "meal_type": "중식",
                    "corner_name": "한식",
                    "menu_name": "자동감지신메뉴",
                    "menu_role": "메인",
                    "source_row_raw": "자동감지신메뉴",
                }
            ]
        },
        headers=AUTH_HEADERS,
    )
    assert resp.status_code == 200

    resp = client.get("/api/dashboard/menu-highlights")
    assert "자동감지신메뉴" in {r["menu_name"] for r in resp.json()["new_menus"]}

    resp = client.put(
        "/api/analysis/menus/new-menu-status", json={"menu_name": "자동감지신메뉴", "is_new": False}
    )
    assert resp.status_code == 200

    resp = client.get("/api/dashboard/menu-highlights")
    assert "자동감지신메뉴" not in {r["menu_name"] for r in resp.json()["new_menus"]}


def test_weekly_menu_predicted_impact_returns_prediction_and_fallback_comment(client):
    _ingest_weekly_menu(client)  # 제육볶음(메인)/계란후라이(부찬), 한식, MONDAY
    for i in range(3):
        _ingest_meal_log(client, f"P{i}", "맛남", menu_name="제육볶음", corner_name="한식")

    resp = client.get(
        "/api/analysis/weekly-menu", params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()}
    )
    assert resp.status_code == 200
    slot = next(s for s in resp.json() if s["corner_name"] == "한식")
    main_plan_id = slot["main"]["plan_id"]
    side_plan_id = slot["sides"][0]["plan_id"]

    resp = client.get(f"/api/analysis/weekly-menu/{main_plan_id}/predicted-impact")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["main_menu"]["menu_name"] == "제육볶음"
    assert isinstance(body["prediction"]["predicted_headcount"], (int, float))
    assert isinstance(body["prediction"]["predicted_share"], (int, float))
    # 이 픽스처엔 과거 처리량 데이터가 없어 None(계산은 되지만 근거가 없다는 뜻)
    assert body["prediction"]["expected_wait_minutes"] is None
    assert "사내 LLM 미설정" in body["summary_comment"]  # 테스트 환경엔 사내 LLM 미설정

    # 메인이 아닌(부찬) plan_id로는 예측을 못 만든다
    resp = client.get(f"/api/analysis/weekly-menu/{side_plan_id}/predicted-impact")
    assert resp.status_code == 404

    resp = client.get("/api/analysis/weekly-menu/999999/predicted-impact")
    assert resp.status_code == 404


def test_weekly_menu_predicted_impact_summary_returns_numbers_without_llm_call(client):
    _ingest_weekly_menu(client)  # 제육볶음(메인)/계란후라이(부찬), 한식, MONDAY
    for i in range(3):
        _ingest_meal_log(client, f"P{i}", "맛남", menu_name="제육볶음", corner_name="한식")

    resp = client.get(
        "/api/analysis/weekly-menu/predicted-impact-summary",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200, resp.text
    rows = resp.json()
    assert len(rows) == 1  # 이 기간엔 메인메뉴 슬롯이 하나(제육볶음)뿐
    row = rows[0]
    assert row["menu_name"] == "제육볶음"
    assert row["corner_name"] == "한식"
    assert row["plan_date"] == MONDAY.isoformat()
    assert row["meal_type"] == "중식"
    assert isinstance(row["prediction"]["predicted_headcount"], (int, float))
    assert isinstance(row["prediction"]["predicted_share"], (int, float))
    assert "expected_wait_minutes" in row["prediction"]
    assert "summary_comment" not in row  # LLM 호출 없이 숫자만


def test_predicted_impact_computes_expected_wait_minutes_from_peak_time_throughput(client):
    _ingest_weekly_menu(client)  # 제육볶음(메인)/계란후라이(부찬), 한식, MONDAY
    # 과거 두 날짜에 피크타임(11:40~12:20, _ingest_meal_log 기본 취식시각 11:52)
    # 취식 기록을 남겨 이 메뉴의 처리량 데이터를 만든다(min_day_count=2 충족).
    for offset_days in (14, 7):
        eaten_date = MONDAY - dt.timedelta(days=offset_days)
        for i in range(3):
            _ingest_meal_log(
                client, f"T{offset_days}_{i}", "맛남", eaten_date=eaten_date, menu_name="제육볶음", corner_name="한식"
            )
    # _baseline_headcount는 daily_corner_stats(집계 테이블)를 보므로, 예상
    # 식수가 0이 아니려면 그 두 날짜분을 집계해둬야 한다.
    resp = client.post(
        "/api/analysis/daily-stats/recompute",
        params={"period_start": (MONDAY - dt.timedelta(days=14)).isoformat(), "period_end": (MONDAY - dt.timedelta(days=7)).isoformat()},
    )
    assert resp.status_code == 200

    resp = client.get(
        "/api/analysis/weekly-menu", params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()}
    )
    main_plan_id = next(s for s in resp.json() if s["corner_name"] == "한식")["main"]["plan_id"]

    resp = client.get(f"/api/analysis/weekly-menu/{main_plan_id}/predicted-impact")
    assert resp.status_code == 200, resp.text
    # 이 픽스처는 과거 실적을 그대로 재현하는 시나리오(배수=1)라 피크 용량을
    # 안 넘어 0이 맞다 — 초과분(overflow>0)이 실제로 나오는 경우는
    # compute_expected_wait_minutes 순수함수 테스트(test_weekly_menu_prediction.py)로
    # 정확히 고정한다. 여기서는 배선(0 이상의 숫자가 나오는지)만 확인.
    expected_wait = resp.json()["prediction"]["expected_wait_minutes"]
    assert expected_wait is not None
    assert expected_wait >= 0


def test_predicted_impact_uses_holiday_history_for_holiday_plan_date(client):
    """회귀 테스트: weekly_menu_prediction.py::compute_predicted_numbers가
    _baseline_headcount에 bool(is_holiday)을 넘기던 버그(2026-08 발견) —
    DayClassification과 비교가 항상 False로 평가돼 주말/공휴일 계획도 매번
    평일 이력만으로 baseline을 계산했다. 평일 이력(식수 낮음)과 주말 이력
    (식수 높음)을 뚜렷이 다르게 시딩해, 실제 계획일이 토요일(주말+공휴일)
    이면 주말 이력이 쓰이는지 확인한다."""
    weekday_history_date = dt.date(2026, 7, 13)  # 월요일
    holiday_history_date = dt.date(2026, 7, 18)  # 토요일
    plan_date = dt.date(2026, 7, 25)  # 토요일 — 이 날짜의 예측이 검증 대상

    for i in range(2):
        _ingest_meal_log(client, f"W{i}", "맛남", eaten_date=weekday_history_date, menu_name="제육볶음")
    for i in range(20):
        _ingest_meal_log(client, f"H{i}", "맛남", eaten_date=holiday_history_date, menu_name="제육볶음")

    resp = client.post(
        "/api/analysis/daily-stats/recompute",
        params={"period_start": weekday_history_date.isoformat(), "period_end": holiday_history_date.isoformat()},
    )
    assert resp.status_code == 200, resp.text

    rows = [
        {
            "plan_date": plan_date.isoformat(),
            "meal_type": "중식",
            "corner_name": "한식",
            "menu_name": "제육볶음",
            "menu_role": "메인",
            "source_row_raw": "제육볶음",
        }
    ]
    resp = client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text

    resp = client.get(
        "/api/analysis/weekly-menu", params={"period_start": plan_date.isoformat(), "period_end": plan_date.isoformat()}
    )
    main_plan_id = next(s for s in resp.json() if s["corner_name"] == "한식")["main"]["plan_id"]

    resp = client.get(f"/api/analysis/weekly-menu/{main_plan_id}/predicted-impact")
    assert resp.status_code == 200, resp.text
    predicted_headcount = resp.json()["prediction"]["predicted_headcount"]
    # 버그가 있으면 평일 이력(2명)만 써서 5명 미만이 나온다 — 주말 이력(20명)을
    # 정상적으로 썼다면 그 근방(배수 적용 후에도 두 자릿수)이어야 한다.
    assert predicted_headcount >= 10

    resp = client.get(
        "/api/analysis/weekly-menu/predicted-impact-summary",
        params={"period_start": plan_date.isoformat(), "period_end": plan_date.isoformat()},
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()[0]["prediction"]["predicted_headcount"] >= 10


def test_new_menu_status_unknown_menu_name_404s(client):
    resp = client.put(
        "/api/analysis/menus/new-menu-status", json={"menu_name": "존재안함", "is_new": True}
    )
    assert resp.status_code == 404


def test_improvement_points_surfaces_congestion_satisfaction_voe(client):
    def eat(employee_id, corner_name, minute, taste="맛남", menu_name=None, comment=None, eaten_date=MONDAY):
        client.post(
            "/api/ingest/meal-log",
            json={
                "rows": [
                    {
                        "eaten_at": dt.datetime.combine(eaten_date, dt.time(11, minute)).isoformat(),
                        "employee_id": employee_id,
                        "meal_type": "중식",
                        "corner_name": corner_name,
                        "menu_name": menu_name,
                        "taste_score": taste,
                        "comment": comment,
                    }
                ]
            },
            headers=AUTH_HEADERS,
        )

    # 혼잡도: 한식은 식수는 많지만(20명) 피크타임(11:40~12:00) 안에는 2명만 —
    # 나머지는 피크 밖(13시)이라 서브속도가 낮게 잡힌다. 일품은 식수는 적지만
    # 전부 피크 안이라 서브속도가 높다.
    for i in range(2):
        eat(f"H{i}", "한식", 45, menu_name="비인기저조메뉴", taste="개선")
    for i in range(18):
        client.post(
            "/api/ingest/meal-log",
            json={
                "rows": [
                    {
                        "eaten_at": dt.datetime.combine(MONDAY, dt.time(13, i % 60)).isoformat(),
                        "employee_id": f"H{i + 2}",
                        "meal_type": "중식",
                        "corner_name": "한식",
                        "menu_name": "비인기저조메뉴",
                        "taste_score": "개선",
                    }
                ]
            },
            headers=AUTH_HEADERS,
        )
    for i in range(5):
        eat(f"I{i}", "일품", 46, menu_name="인기메뉴", taste="맛남")

    resp = client.post(
        "/api/analysis/daily-stats/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200

    resp = client.post(
        "/api/analysis/menu-performance/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200

    # VOE: 이번 달엔 "위생" 코멘트가 늘고, 지난달엔 거의 없었음
    prior_month_date = (MONDAY.replace(day=1) - dt.timedelta(days=1))
    eat("V0", "한식", 50, comment="위생 상태가 별로였어요", eaten_date=prior_month_date)
    for i in range(4):
        eat(f"V{i + 1}", "한식", 51, comment="위생이 너무 안 좋아요", eaten_date=MONDAY)

    resp = client.get(
        "/api/dashboard/improvement-points",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    points = resp.json()
    axes = {p["axis"] for p in points}
    assert "congestion" in axes
    assert "satisfaction" in axes
    assert "voe" in axes

    congestion = next(p for p in points if p["axis"] == "congestion")
    assert "한식" in congestion["title"]
    satisfaction = next(p for p in points if p["axis"] == "satisfaction")
    assert "비인기저조메뉴" in satisfaction["title"]
    voe = next(p for p in points if p["axis"] == "voe")
    assert "위생" in voe["title"]
    # 테스트 환경엔 사내 LLM이 설정돼 있지 않으므로 폴백 요약(원문 예시)이 붙는다 —
    # 건수만이 아니라 실제 코멘트 내용도 함께 보여줘야 한다는 요청(2026-07)에 대응.
    assert "voe_summary" in voe
    assert "위생" in voe["voe_summary"]


def test_list_weekly_menu_groups_main_and_sides_with_deadline(client):
    _ingest_weekly_menu(client)  # 제육볶음(메인)/계란후라이(부찬), 한식, MONDAY

    resp = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    slots = resp.json()
    assert len(slots) == 1
    slot = slots[0]
    assert slot["corner_name"] == "한식"
    assert slot["main"]["menu_name"] == "제육볶음"
    assert slot["main"]["role_source"] == "규칙기반"
    assert [s["menu_name"] for s in slot["sides"]] == ["계란후라이"]
    assert slot["feedback_deadline"] == (MONDAY - dt.timedelta(days=7)).isoformat()
    expected_past_deadline = dt.date.today() > (MONDAY - dt.timedelta(days=7))
    assert slot["is_past_deadline"] == expected_past_deadline


def test_update_weekly_menu_role_locks_role_source_to_manual(client, db_session):
    _ingest_weekly_menu(client)
    resp = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    side_plan_id = resp.json()[0]["sides"][0]["plan_id"]

    # 부찬을 부찬인 채로(역할은 그대로) 관리자가 "확인 완료"로 저장하는 상황을
    # 가정 — role_source만 관리자수동으로 잠기는지 확인(메인 중복 생기는
    # 케이스는 group_weekly_menu_rows의 별도 관심사라 여기서 안 섞음).
    resp = client.put(f"/api/analysis/weekly-menu/{side_plan_id}/role", json={"menu_role": "부찬"})
    assert resp.status_code == 200
    body = resp.json()
    assert body["menu_role"] == "부찬"
    assert body["role_source"] == "관리자수동"

    from app.models.logs import WeeklyMenuPlan

    plan = db_session.get(WeeklyMenuPlan, side_plan_id)
    db_session.refresh(plan)
    assert plan.role_source.value == "관리자수동"


def test_update_weekly_menu_role_to_main_demotes_previous_main(client, db_session):
    # 부찬을 메인으로 승격하면 기존 메인은 자동으로 부찬으로 내려가야 한다 —
    # 안 그러면 슬롯에 메인이 2개 남아 조회 화면과 시뮬레이션이 서로 다른
    # 메뉴를 "그 날의 메인"으로 볼 수 있다(실사용 중 발견).
    _ingest_weekly_menu(client)
    resp = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    slot = resp.json()[0]
    old_main_plan_id = slot["main"]["plan_id"]
    side_plan_id = slot["sides"][0]["plan_id"]

    resp = client.put(f"/api/analysis/weekly-menu/{side_plan_id}/role", json={"menu_role": "메인"})
    assert resp.status_code == 200

    from app.models.logs import WeeklyMenuPlan

    old_main_plan = db_session.get(WeeklyMenuPlan, old_main_plan_id)
    db_session.refresh(old_main_plan)
    assert old_main_plan.menu_role.value == "부찬"
    assert old_main_plan.role_source.value == "관리자수동"

    # 조회 화면과 실제 DB 상태가 일치하는지 — 메인이 정확히 하나만 남아야 함
    resp = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    slot = resp.json()[0]
    assert slot["main"]["plan_id"] == side_plan_id
    assert [s["plan_id"] for s in slot["sides"]] == [old_main_plan_id]


def test_update_weekly_menu_role_unknown_plan_returns_404(client):
    resp = client.put("/api/analysis/weekly-menu/999999/role", json={"menu_role": "메인"})
    assert resp.status_code == 404


def test_weekly_menu_feedback_create_and_list(client, db_session):
    _ingest_weekly_menu(client)
    from app.models.master import CornerMaster

    corner = db_session.query(CornerMaster).filter_by(corner_name="한식").one()

    resp = client.post(
        "/api/analysis/weekly-menu/feedback",
        json={"plan_date": MONDAY.isoformat(), "corner_id": corner.corner_id, "comment": "이 부찬 조합 별로예요"},
    )
    assert resp.status_code == 200
    assert resp.json()["comment"] == "이 부찬 조합 별로예요"

    resp = client.get(
        "/api/analysis/weekly-menu/feedback",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    items = resp.json()
    assert len(items) == 1
    assert items[0]["corner_name"] == "한식"
    assert items[0]["comment"] == "이 부찬 조합 별로예요"


def test_weekly_menu_feedback_unknown_corner_returns_404(client):
    resp = client.post(
        "/api/analysis/weekly-menu/feedback",
        json={"plan_date": MONDAY.isoformat(), "corner_id": 999999, "comment": "테스트"},
    )
    assert resp.status_code == 404


def test_menu_side_combinations_compares_satisfaction_by_combo(client, db_session):
    week1 = MONDAY
    week2 = MONDAY + dt.timedelta(days=7)

    def ingest_slot(plan_date, side_name):
        resp = client.post(
            "/api/ingest/weekly-menu",
            json={
                "rows": [
                    {
                        "plan_date": plan_date.isoformat(),
                        "meal_type": "중식",
                        "corner_name": "한식",
                        "menu_name": "제육볶음",
                        "menu_role": "메인",
                        "source_row_raw": f"제육볶음\n{side_name}",
                    },
                    {
                        "plan_date": plan_date.isoformat(),
                        "meal_type": "중식",
                        "corner_name": "한식",
                        "menu_name": side_name,
                        "menu_role": "부찬",
                        "source_row_raw": f"제육볶음\n{side_name}",
                    },
                ]
            },
            headers=AUTH_HEADERS,
        )
        assert resp.status_code == 200

    ingest_slot(week1, "계란찜")
    ingest_slot(week2, "김치")

    for i in range(3):
        _ingest_meal_log(client, f"W1_{i}", "맛남", eaten_date=week1, menu_name="제육볶음", corner_name="한식")
    for i in range(3):
        _ingest_meal_log(client, f"W2_{i}", "개선", eaten_date=week2, menu_name="제육볶음", corner_name="한식")

    _set_food_vector(db_session, "제육볶음", [0.5] * 10)
    _set_food_vector(db_session, "계란찜", [0.1] * 10)
    _set_food_vector(db_session, "김치", [0.9] * 10)

    resp = client.get(
        "/api/analysis/menu-combinations/제육볶음",
        params={"period_start": week1.isoformat(), "period_end": week2.isoformat()},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["menu_name"] == "제육볶음"
    combos = body["combos"]
    assert len(combos) == 2
    # 맛남(5점)이 개선(1점)보다 만족도가 높으므로 계란찜 조합이 먼저 나와야 함
    assert combos[0]["sides"] == ["계란찜"]
    assert combos[0]["avg_satisfaction"] == 5.0
    assert combos[0]["day_count"] == 1
    assert combos[1]["sides"] == ["김치"]
    assert combos[1]["avg_satisfaction"] == 1.0
    assert combos[0]["nutrition_profile"]["매운맛"] == 0.3  # (0.5+0.1)/2


def test_menu_side_combinations_unknown_menu_returns_404(client):
    resp = client.get(
        "/api/analysis/menu-combinations/존재하지않는메뉴",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 404


def test_reclassify_weekly_menu_roles_endpoint_no_op_when_llm_unconfigured(client):
    _ingest_weekly_menu(client)

    resp = client.post(
        "/api/analysis/weekly-menu/reclassify-roles-with-llm",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    # 사내 LLM 미설정 환경에서는 모의 응답이 "메인: OO" 형식으로 파싱되지 않으므로 0건
    assert resp.json()["reclassified_slots"] == 0


def test_dashboard_weekly_summary_classifies_days(client):
    resp = client.get(
        "/api/dashboard/weekly-summary",
        params={"start_date": MONDAY.isoformat(), "end_date": (MONDAY + dt.timedelta(days=6)).isoformat()},
    )
    assert resp.status_code == 200
    days = resp.json()
    assert len(days) == 7
    saturday = next(d for d in days if d["date"] == (MONDAY + dt.timedelta(days=5)).isoformat())
    assert saturday["classification"] == "주말+공휴일"


def test_corner_analysis_requires_daily_stats(client, db_session):
    _ingest_weekly_menu(client)
    _ingest_meal_log(client, "E1", "맛남")

    from app.services.aggregation import aggregate_daily_stats

    aggregate_daily_stats(db_session, MONDAY)

    resp = client.get(
        "/api/analysis/corners",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    rows = resp.json()
    hansik = next(r for r in rows if r["corner_name"] == "한식")
    assert hansik["headcount_total"] == 1


def test_corner_analysis_merges_take_out_aliases_and_excludes_on_request(client, db_session):
    from app.services.aggregation import aggregate_daily_stats

    _ingest_weekly_menu(client)
    _ingest_meal_log(client, "E1", "맛남", corner_name="Take Out R")
    _ingest_meal_log(client, "E2", "맛남", corner_name="Take Out M")
    _ingest_meal_log(client, "E3", "맛남", corner_name="Take Out L")
    aggregate_daily_stats(db_session, MONDAY)

    resp = client.get(
        "/api/analysis/corners",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    rows = resp.json()
    take_out_rows = [r for r in rows if r["corner_name"] == "Take Out"]
    assert len(take_out_rows) == 1  # R/M/L 세 이름이 하나로 합쳐짐
    assert take_out_rows[0]["headcount_total"] == 3
    assert not any(r["corner_name"] in ("Take Out R", "Take Out M", "Take Out L") for r in rows)

    resp = client.get(
        "/api/analysis/corners",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "exclude_take_out": True},
    )
    rows = resp.json()
    assert not any(r["corner_name"] == "Take Out" for r in rows)


def test_corner_analysis_sorts_green_meat_last_regardless_of_headcount(client, db_session):
    from app.services.aggregation import aggregate_daily_stats

    _ingest_weekly_menu(client)
    _ingest_meal_log(client, "E1", "맛남", corner_name="한식")
    for i in range(5):
        _ingest_meal_log(client, f"G{i}", "맛남", corner_name="그린미트")
    aggregate_daily_stats(db_session, MONDAY)

    resp = client.get(
        "/api/analysis/corners",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    rows = resp.json()
    # 그린미트가 식수는 제일 많아도(5건 vs 1건) 항상 마지막이어야 한다
    assert rows[-1]["corner_name"] == "그린미트"
    assert rows[-1]["headcount_total"] == 5


def test_corner_analysis_trend_groups_by_period_and_corner(client, db_session):
    from app.services.aggregation import aggregate_daily_stats

    _ingest_weekly_menu(client)
    _ingest_meal_log(client, "E1", "맛남", eaten_date=MONDAY)
    _ingest_meal_log(client, "E2", "맛남", eaten_date=MONDAY + dt.timedelta(days=7))
    aggregate_daily_stats(db_session, MONDAY)
    aggregate_daily_stats(db_session, MONDAY + dt.timedelta(days=7))

    resp = client.get(
        "/api/analysis/corners/trend",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": (MONDAY + dt.timedelta(days=7)).isoformat(),
            "granularity": "weekly",
        },
    )
    assert resp.status_code == 200
    rows = resp.json()
    hansik_rows = [r for r in rows if r["corner_name"] == "한식"]
    assert len(hansik_rows) == 2  # 서로 다른 주 2개
    assert {r["headcount"] for r in hansik_rows} == {1, 1}


def test_daily_stats_recompute_backfills_range_for_corner_and_home_views(client):
    # 과거 기간(예: 6개월치)을 한꺼번에 적재했을 때, 매일 새벽 스케줄러가 "어제"
    # 하루치만 계산하는 것과 달리 이 엔드포인트는 기간 전체를 한 번에 채워야 한다.
    _ingest_weekly_menu(client)
    _ingest_meal_log(client, "E1", "맛남", eaten_date=MONDAY)
    _ingest_meal_log(client, "E2", "맛남", eaten_date=MONDAY + dt.timedelta(days=1))

    resp = client.post(
        "/api/analysis/daily-stats/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=1)).isoformat()},
    )
    assert resp.status_code == 200
    assert resp.json()["days_processed"] == 2

    corners_resp = client.get(
        "/api/analysis/corners",
        params={"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=1)).isoformat()},
    )
    hansik = next(r for r in corners_resp.json() if r["corner_name"] == "한식")
    assert hansik["headcount_total"] == 2

    divisions_resp = client.get(
        "/api/analysis/divisions",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": (MONDAY + dt.timedelta(days=1)).isoformat(),
            "granularity": "daily",
        },
    )
    assert sum(r["headcount"] for r in divisions_resp.json()) == 2


def test_daily_stats_recompute_rejects_inverted_range(client):
    resp = client.post(
        "/api/analysis/daily-stats/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": (MONDAY - dt.timedelta(days=1)).isoformat()},
    )
    assert resp.status_code == 400


def test_chat_stream_falls_back_to_mock_when_llm_unconfigured(client):
    resp = client.post(
        "/api/chat/stream", json={"messages": [{"role": "user", "content": "이번주 만족도 어때?"}]}
    )
    assert resp.status_code == 200
    # 모의 응답은 단어 단위 SSE 청크로 쪼개져 오므로 한 단어로 확인한다.
    assert "미설정" in resp.text
    assert "data: [DONE]" in resp.text


def test_weekly_summary_export_returns_xlsx(client):
    resp = client.get(
        "/api/dashboard/weekly-summary/export",
        params={"start_date": MONDAY.isoformat(), "end_date": (MONDAY + dt.timedelta(days=6)).isoformat()},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert resp.content[:2] == b"PK"  # xlsx는 zip 컨테이너이므로 PK 매직 바이트로 시작


def test_meal_log_export_returns_xlsx_for_selected_period(client, db_session):
    _ingest_weekly_menu(client)
    _ingest_meal_log(client, "E11111", "맛남", "맛있어요", company_name="삼성전자")
    _ingest_meal_log(client, "E22222", "보통", None, company_name="삼성SDI")
    # 기간 밖 데이터는 제외돼야 함
    _ingest_meal_log(client, "E33333", "개선", None, eaten_date=MONDAY + dt.timedelta(days=10))

    resp = client.get(
        "/api/dashboard/meal-log/export",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert resp.content[:2] == b"PK"

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content))
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("취식일시", "사번", "구분", "회사명", "식사구분", "코너", "메뉴", "맛평가", "의견")
    assert len(rows) == 3  # header + 2 rows (기간 밖 1건 제외)
    employee_ids = {row[1] for row in rows[1:]}
    assert employee_ids == {"E11111", "E22222"}


def test_simulation_what_if_returns_all_corners(client):
    _ingest_weekly_menu(client)
    resp = client.post(
        "/api/simulation/what-if",
        json={"target_date": MONDAY.isoformat(), "meal_type": "중식", "weather": "비"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["classification"] == "평일"
    assert any(c["corner_name"] == "한식" for c in body["corners"])


def test_congestion_forecast_adjusts_for_planned_menu_popularity(client, db_session):
    def eat(employee_id, corner_name, menu_name):
        r = client.post(
            "/api/ingest/meal-log",
            json={
                "rows": [
                    {
                        "eaten_at": dt.datetime.combine(MONDAY, dt.time(11, 50)).isoformat(),
                        "employee_id": employee_id,
                        "meal_type": "중식",
                        "corner_name": corner_name,
                        "menu_name": menu_name,
                        "taste_score": "맛남",
                    }
                ]
            },
            headers=AUTH_HEADERS,
        )
        assert r.status_code == 200

    # 한식: 인기메뉴A(8명) vs 비인기메뉴B(2명), 일품: 메뉴C(10명) — 전체 20명
    # share(A)=8/20=0.4, share(B)=2/20=0.1, 한식 평균share=(0.4+0.1)/2=0.25
    # → A가 계획되면 배수 0.4/0.25=1.6
    for i in range(8):
        eat(f"A{i}", "한식", "인기메뉴A")
    for i in range(2):
        eat(f"B{i}", "한식", "비인기메뉴B")
    for i in range(10):
        eat(f"C{i}", "일품", "메뉴C")

    resp = client.post(
        "/api/analysis/daily-stats/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    resp = client.post(
        "/api/analysis/menu-performance/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200

    from app.models.master import MenuMaster

    menu_a = db_session.query(MenuMaster).filter_by(menu_name="인기메뉴A").one()

    # 미래 평일(다음 주 화요일)에 인기메뉴A가 한식 코너 메인으로 계획됨
    future_date = MONDAY + dt.timedelta(days=8)
    resp = client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                {
                    "plan_date": future_date.isoformat(),
                    "meal_type": "중식",
                    "corner_name": "한식",
                    "menu_name": "인기메뉴A",
                    "menu_role": "메인",
                    "source_row_raw": "인기메뉴A",
                }
            ]
        },
        headers=AUTH_HEADERS,
    )
    assert resp.status_code == 200

    resp = client.get(
        "/api/simulation/congestion-forecast",
        params={"target_date": future_date.isoformat(), "meal_type": "중식"},
    )
    assert resp.status_code == 200
    hansik = next(c for c in resp.json()["corners"] if c["corner_name"] == "한식")
    assert hansik["planned_menu_id"] == menu_a.menu_id
    assert hansik["menu_popularity_multiplier"] == 1.6
    # 코너 baseline(A+B 합계 10명) * 1.6배 = 16.0
    assert hansik["predicted_headcount"] == 16.0
    # 실측 meal_log가 전부 피크타임(11:50, 11:40~12:20) 안에서만 찍혀
    # peak_share_ratio=1.0 — 예상 피크 식수는 예상 식수와 같다("최고 혼잡
    # 예상 코너" 카드용 신규 필드, 2026-07).
    assert hansik["expected_peak_headcount"] == 16.0

    ilpum = next(c for c in resp.json()["corners"] if c["corner_name"] == "일품")
    assert ilpum["planned_menu_id"] is None
    assert ilpum["menu_popularity_multiplier"] is None
    assert ilpum["expected_peak_headcount"] == ilpum["predicted_headcount"]


def test_what_if_uses_quadrant_multiplier_for_planned_menu_with_performance_data(client, db_session):
    def eat(employee_id, corner_name, menu_name):
        r = client.post(
            "/api/ingest/meal-log",
            json={
                "rows": [
                    {
                        "eaten_at": dt.datetime.combine(MONDAY, dt.time(11, 50)).isoformat(),
                        "employee_id": employee_id,
                        "meal_type": "중식",
                        "corner_name": corner_name,
                        "menu_name": menu_name,
                        "taste_score": "맛남",
                    }
                ]
            },
            headers=AUTH_HEADERS,
        )
        assert r.status_code == 200

    # 수요 median=10, 평가건수 10 미만이면 표본부족으로 빠지므로 A=12(고수요,
    # 평가 충분)가 "인기메뉴"(POPULAR, 배수 1.20)로 분류되게 구성.
    for i in range(12):
        eat(f"A{i}", "한식", "인기메뉴A")
    for i in range(2):
        eat(f"B{i}", "한식", "비인기메뉴B")
    for i in range(10):
        eat(f"C{i}", "일품", "메뉴C")

    resp = client.post(
        "/api/analysis/daily-stats/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    resp = client.post(
        "/api/analysis/menu-performance/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200

    from app.models.master import CornerMaster, MenuMaster

    menu_a = db_session.query(MenuMaster).filter_by(menu_name="인기메뉴A").one()
    hansik_corner = db_session.query(CornerMaster).filter_by(corner_name="한식").one()

    # 확인: A가 실제로 인기메뉴(POPULAR)로 분류됐는지
    resp = client.get(
        "/api/analysis/menu-performance",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    a_row = next(r for r in resp.json() if r["menu_name"] == "인기메뉴A")
    assert a_row["quadrant"] == "인기메뉴"

    future_date = MONDAY + dt.timedelta(days=8)
    resp = client.post(
        "/api/simulation/what-if",
        json={
            "target_date": future_date.isoformat(),
            "meal_type": "중식",
            "new_menu_corner_id": hansik_corner.corner_id,
            "planned_menu_id": menu_a.menu_id,
        },
    )
    assert resp.status_code == 200
    hansik = next(c for c in resp.json()["corners"] if c["corner_name"] == "한식")
    # baseline(A+B 합계 14명) * 1.20(인기메뉴 배수, 기본 1.15보다 큼) = 16.8
    assert hansik["baseline_headcount"] == 14.0
    assert hansik["predicted_headcount"] == 16.8


def test_meal_log_ingest_classifies_division_from_company_name(client, db_session):
    _ingest_meal_log(client, "E1001", "맛남", company_name="삼성전자")
    _ingest_meal_log(client, "E1002", "맛남", company_name="삼성SDI")
    _ingest_meal_log(client, "E1003", "맛남", company_name="지리산")
    _ingest_meal_log(client, "E1004", "맛남")  # company_name 없음(과거 방식 호환)

    from app.models.master import EmployeeMaster

    employees = {
        e.employee_id: e for e in db_session.query(EmployeeMaster).filter(
            EmployeeMaster.employee_id.in_(["E1001", "E1002", "E1003", "E1004"])
        )
    }
    assert employees["E1001"].division.value == "본사"
    assert employees["E1001"].company_name == "삼성전자"
    assert employees["E1002"].division.value == "계열사"
    assert employees["E1002"].company_name == "삼성SDI"
    assert employees["E1003"].division.value == "기타"
    assert employees["E1004"].division.value == "기타"
    assert employees["E1004"].company_name is None


def test_division_analysis_daily_breakdown(client, db_session):
    _ingest_meal_log(client, "E3001", "맛남", company_name="삼성전자")
    _ingest_meal_log(client, "E3002", "맛남", company_name="삼성SDI")
    _ingest_meal_log(client, "E3003", "맛남", company_name="지리산")

    from app.services.aggregation import aggregate_daily_stats

    aggregate_daily_stats(db_session, MONDAY)

    resp = client.get(
        "/api/analysis/divisions",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    by_division = {r["division"]: r["headcount"] for r in resp.json()}
    assert by_division == {"본사": 1, "계열사": 1, "기타": 1}


def test_division_analysis_monthly_granularity_combines_days(client, db_session):
    from app.services.aggregation import aggregate_daily_stats

    day1 = MONDAY
    day2 = MONDAY + dt.timedelta(days=1)
    _ingest_meal_log(client, "E4001", "맛남", company_name="삼성전자", eaten_date=day1)
    _ingest_meal_log(client, "E4002", "맛남", company_name="삼성전자", eaten_date=day2)
    aggregate_daily_stats(db_session, day1)
    aggregate_daily_stats(db_session, day2)

    resp = client.get(
        "/api/analysis/divisions",
        params={"period_start": day1.isoformat(), "period_end": day2.isoformat(), "granularity": "monthly"},
    )
    assert resp.status_code == 200
    rows = resp.json()
    assert len(rows) == 1  # 같은 달 두 날짜가 한 버킷으로 합쳐져야 함
    assert rows[0]["period"] == day1.strftime("%Y-%m")
    assert rows[0]["headcount"] == 2


def test_division_analysis_classification_filter(client, db_session):
    from app.services.aggregation import aggregate_daily_stats

    saturday = MONDAY + dt.timedelta(days=5)
    _ingest_meal_log(client, "E5001", "맛남", company_name="삼성전자", eaten_date=MONDAY)
    _ingest_meal_log(client, "E5002", "맛남", company_name="삼성전자", eaten_date=saturday)
    aggregate_daily_stats(db_session, MONDAY)
    aggregate_daily_stats(db_session, saturday)

    resp = client.get(
        "/api/analysis/divisions",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": saturday.isoformat(),
            "classification": "주말+공휴일",
        },
    )
    rows = resp.json()
    assert len(rows) == 1
    assert rows[0]["period"] == saturday.isoformat()


def test_division_analysis_family_day_classification_filter(client, db_session):
    from app.services.aggregation import aggregate_daily_stats
    from app.services.holidays import family_day_of_month

    family_day = family_day_of_month(MONDAY.year, MONDAY.month)
    assert family_day == MONDAY + dt.timedelta(days=4)  # 2026-07-24(금) — MONDAY가 속한 주와 같은 주

    _ingest_meal_log(client, "E5001", "맛남", company_name="삼성전자", eaten_date=MONDAY)
    _ingest_meal_log(client, "E5002", "맛남", company_name="삼성전자", eaten_date=family_day)
    aggregate_daily_stats(db_session, MONDAY)
    aggregate_daily_stats(db_session, family_day)

    resp = client.get(
        "/api/analysis/divisions",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": family_day.isoformat(),
            "classification": "패밀리데이",
        },
    )
    rows = resp.json()
    assert len(rows) == 1
    assert rows[0]["period"] == family_day.isoformat()

    # 패밀리데이는 더는 "평일" 버킷에 안 섞여야 한다
    weekday_resp = client.get(
        "/api/analysis/divisions",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": family_day.isoformat(),
            "classification": "평일",
        },
    )
    weekday_rows = weekday_resp.json()
    assert len(weekday_rows) == 1
    assert weekday_rows[0]["period"] == MONDAY.isoformat()


def _set_food_vector(db_session, menu_name: str, vector: list[float]):
    from app.models.master import MenuMaster

    menu = db_session.query(MenuMaster).filter_by(menu_name=menu_name).one()
    menu.food_vector = vector
    db_session.commit()


def test_taste_clusters_recompute_and_list(client, db_session):
    # 매운맛 그룹(spicy 높음) 4명, 순한맛 그룹(protein 높음) 4명 — 명확히 갈리게 구성
    for i in range(4):
        _ingest_meal_log(client, f"S{i}", "맛남", company_name=None)
        r = client.post(
            "/api/ingest/meal-log",
            json={
                "rows": [
                    {
                        "eaten_at": dt.datetime.combine(MONDAY, dt.time(12, i)).isoformat(),
                        "employee_id": f"S{i}",
                        "meal_type": "중식",
                        "corner_name": "한식",
                        "menu_name": "매운메뉴",
                        "taste_score": "맛남",
                    }
                ]
            },
            headers=AUTH_HEADERS,
        )
        assert r.status_code == 200
    for i in range(4):
        client.post(
            "/api/ingest/meal-log",
            json={
                "rows": [
                    {
                        "eaten_at": dt.datetime.combine(MONDAY, dt.time(12, 30 + i)).isoformat(),
                        "employee_id": f"P{i}",
                        "meal_type": "중식",
                        "corner_name": "그린미트",
                        "menu_name": "고단백메뉴",
                        "taste_score": "맛남",
                    }
                ]
            },
            headers=AUTH_HEADERS,
        )

    _set_food_vector(db_session, "매운메뉴", [0.9, 0.1, 0.1, 0.1, 0.1, 0.2, 0.1, 0.1, 0.1, 0.1])
    _set_food_vector(db_session, "고단백메뉴", [0.1, 0.1, 0.1, 0.1, 0.1, 0.9, 0.1, 0.1, 0.1, 0.1])

    resp = client.post("/api/analysis/users/taste-profile/recompute")
    assert resp.status_code == 200
    assert resp.json()["updated_employees"] == 8

    # 표본 부족(8명인데 k=10 요구 -> 최소 20명 필요)이면 400
    resp = client.post("/api/analysis/users/taste-clusters/recompute", params={"k": 10})
    assert resp.status_code == 400

    resp = client.post("/api/analysis/users/taste-clusters/recompute", params={"k": 2})
    assert resp.status_code == 200
    assert resp.json()["clusters_created"] == 2

    resp = client.get("/api/analysis/users/taste-clusters")
    assert resp.status_code == 200
    clusters = resp.json()
    assert len(clusters) == 2
    assert {c["size"] for c in clusters} == {4, 4}
    labels = {c["label"] for c in clusters}
    assert any("매운맛" in label for label in labels)
    assert any("단백질" in label for label in labels)

    # 개별 사번 조회에도 소속 군집 라벨이 붙는지 확인
    resp = client.get("/api/analysis/users/S0/taste-profile")
    assert resp.status_code == 200
    assert resp.json()["cluster_label"] is not None


def test_taste_clusters_exclude_take_out_from_dominant_corner_and_top_menus(client, db_session):
    # 사번당 Take Out 방문(2회)이 한식 방문(1회)보다 많게 구성 — 제외 규칙이 없으면
    # dominant_corner/top_menus가 "Take Out"/"선택형 Take out"으로 잘못 나온다.
    def eat(employee_id, corner_name, menu_name, minute):
        r = client.post(
            "/api/ingest/meal-log",
            json={
                "rows": [
                    {
                        "eaten_at": dt.datetime.combine(MONDAY, dt.time(11, minute)).isoformat(),
                        "employee_id": employee_id,
                        "meal_type": "중식",
                        "corner_name": corner_name,
                        "menu_name": menu_name,
                        "taste_score": "맛남",
                    }
                ]
            },
            headers=AUTH_HEADERS,
        )
        assert r.status_code == 200

    for i in range(4):
        eat(f"T{i}", "한식", "김치찌개", i)
        eat(f"T{i}", "Take Out", "선택형 Take out", 10 + i)
        eat(f"T{i}", "Take Out", "선택형 Take out", 20 + i)

    _set_food_vector(db_session, "김치찌개", [0.5] * 10)

    resp = client.post("/api/analysis/users/taste-profile/recompute")
    assert resp.status_code == 200
    assert resp.json()["updated_employees"] == 4

    resp = client.post("/api/analysis/users/taste-clusters/recompute", params={"k": 1})
    assert resp.status_code == 200
    assert resp.json()["clusters_created"] == 1

    resp = client.get("/api/analysis/users/taste-clusters")
    cluster = resp.json()[0]
    assert cluster["dominant_corner"] == "한식"
    assert "선택형 Take out" not in cluster["top_menus"]
    assert cluster["top_menus"] == ["김치찌개"]


def test_menu_affinity_finds_co_occurring_menu(client):
    # 떡볶이 먹는 3명 중 2명이 짜장면도 먹음
    def eat(employee_id, menu_name, minute):
        client.post(
            "/api/ingest/meal-log",
            json={
                "rows": [
                    {
                        "eaten_at": dt.datetime.combine(MONDAY, dt.time(11, minute)).isoformat(),
                        "employee_id": employee_id,
                        "meal_type": "중식",
                        "corner_name": "분식",
                        "menu_name": menu_name,
                        "taste_score": "맛남",
                    }
                ]
            },
            headers=AUTH_HEADERS,
        )

    for emp in ["A1", "A2", "A3"]:
        eat(emp, "떡볶이", 0)
    for emp in ["A1", "A2"]:
        eat(emp, "짜장면", 10)
    for emp in ["B1", "B2", "B3"]:
        eat(emp, "돈까스", 20)

    resp = client.get(
        "/api/analysis/menu-affinity/떡볶이",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "min_co_count": 1},
    )
    assert resp.status_code == 200
    rows = resp.json()
    jjajang = next(r for r in rows if r["menu_name"] == "짜장면")
    assert jjajang["co_count"] == 2
    assert jjajang["lift"] > 1


def test_menu_affinity_unknown_menu_returns_404(client):
    resp = client.get(
        "/api/analysis/menu-affinity/존재하지않는메뉴",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 404


def test_corner_core_layer_summary_compares_all_corners_in_one_call(client):
    # E1~E3: 한식 코어층(4번 방문, share=1.0). N1~N3: 한식은 1번만(미달),
    # 양식을 5번 방문(share 5/6)해 양식 코어층.
    for emp in ["E1", "E2", "E3"]:
        for day_offset, menu in enumerate(["떡볶이", "짜장면", "떡볶이", "짜장면"]):
            _ingest_meal_log(
                client,
                emp,
                "맛남",
                eaten_date=MONDAY + dt.timedelta(days=day_offset),
                menu_name=menu,
                corner_name="한식",
            )
    for emp in ["N1", "N2", "N3"]:
        _ingest_meal_log(client, emp, "맛남", eaten_date=MONDAY, menu_name="김치찌개", corner_name="한식")
        for day_offset, menu in enumerate(["스테이크", "파스타", "스테이크", "파스타", "스테이크"]):
            _ingest_meal_log(
                client,
                emp,
                "맛남",
                eaten_date=MONDAY + dt.timedelta(days=day_offset + 1),
                menu_name=menu,
                corner_name="양식",
            )

    resp = client.get(
        "/api/analysis/corners/core-layer-summary",
        params={"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=10)).isoformat()},
    )
    assert resp.status_code == 200
    body = resp.json()
    by_name = {r["corner_name"]: r for r in body}
    assert by_name["한식"]["core_employee_count"] == 3
    assert by_name["한식"]["non_core_employee_count"] == 3
    assert by_name["양식"]["core_employee_count"] == 3
    assert by_name["양식"]["non_core_employee_count"] == 3


def test_corner_core_layer_summary_excludes_take_out(client):
    # E1~E3: 한식 3번(코어층 판정에 딱 충분, share=1.0 — Take Out 방문이 섞이지
    # 않았을 때 기준) + Take Out 10번. Take Out을 분모에서 안 빼면
    # share=3/13≈0.23으로 min_share(0.3) 미달돼 코어층에서 빠지게 된다 —
    # Take Out을 완전히 제외해야 원래 의도대로 코어층(share=1.0)으로 잡힌다.
    for emp in ["E1", "E2", "E3"]:
        for day_offset, menu in enumerate(["떡볶이", "짜장면", "떡볶이"]):
            _ingest_meal_log(
                client,
                emp,
                "맛남",
                eaten_date=MONDAY + dt.timedelta(days=day_offset),
                menu_name=menu,
                corner_name="한식",
            )
        for day_offset in range(10):
            _ingest_meal_log(
                client,
                emp,
                "맛남",
                eaten_date=MONDAY + dt.timedelta(days=day_offset),
                menu_name="테이크아웃메뉴",
                corner_name="Take Out",
            )

    resp = client.get(
        "/api/analysis/corners/core-layer-summary",
        params={"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=10)).isoformat()},
    )
    assert resp.status_code == 200
    body = resp.json()
    corner_names = {r["corner_name"] for r in body}
    assert "Take Out" not in corner_names  # Take Out 자체는 요약 표에서 빠진다

    by_name = {r["corner_name"]: r for r in body}
    assert by_name["한식"]["core_employee_count"] == 3  # Take Out이 분모를 오염시키지 않음


def test_corner_menu_throughput_sorts_slowest_menu_first(client, db_session):
    # 느린메뉴가 나온 날은 피크타임(11:40~12:00, _ingest_meal_log 기본 취식시각
    # 11:52는 이 구간 안)에 2명만 먹고, 빠른메뉴가 나온 날은 10명이 먹는다 —
    # 처리량(분당 서브)이 낮을수록(=느릴수록) 먼저 나와야 한다.
    for i in range(2):
        _ingest_meal_log(client, f"S{i}", "맛남", eaten_date=MONDAY, menu_name="느린메뉴", corner_name="한식")
    for i in range(2):
        _ingest_meal_log(
            client, f"S{i}b", "맛남", eaten_date=MONDAY + dt.timedelta(days=7), menu_name="느린메뉴", corner_name="한식"
        )
    for i in range(10):
        _ingest_meal_log(
            client, f"F{i}", "맛남", eaten_date=MONDAY + dt.timedelta(days=1), menu_name="빠른메뉴", corner_name="한식"
        )
    for i in range(10):
        _ingest_meal_log(
            client, f"F{i}b", "맛남", eaten_date=MONDAY + dt.timedelta(days=8), menu_name="빠른메뉴", corner_name="한식"
        )

    from app.models.master import CornerMaster

    corner = db_session.query(CornerMaster).filter_by(corner_name="한식").one()

    resp = client.get(
        f"/api/analysis/corners/{corner.corner_id}/menu-throughput",
        params={"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=8)).isoformat()},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert [m["menu_name"] for m in body["menus"]] == ["느린메뉴", "빠른메뉴"]
    assert body["menus"][0]["avg_throughput"] < body["menus"][1]["avg_throughput"]
    assert body["menus"][0]["day_count"] == 2
    assert body["overall_avg_throughput"] is not None


def test_corner_menu_throughput_unknown_corner_returns_404(client):
    resp = client.get(
        "/api/analysis/corners/999999/menu-throughput",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 404


def test_voe_by_category_groups_comments_into_fixed_categories(client):
    _ingest_meal_log(client, "E1", "맛남", comment="정말 맛있어요")
    _ingest_meal_log(client, "E2", "개선", comment="국이 너무 싱거워요")
    _ingest_meal_log(client, "E3", "개선", comment="위생 상태가 별로였어요")
    _ingest_meal_log(client, "E4", "개선", comment="직원분이 불친절했어요")
    _ingest_meal_log(client, "E5", "보통", comment="그냥 평범했어요")  # 기타로 분류
    _ingest_meal_log(client, "E6", "맛남", comment=None)  # 코멘트 없음 — 집계 제외

    resp = client.get("/api/dashboard/voe-by-category", params={"period": f"{MONDAY.isoformat()[:7]}-01"})
    assert resp.status_code == 200
    body = resp.json()

    assert body["total_comments"] == 5
    categories = {c["category"]: c for c in body["categories"]}
    assert [c["category"] for c in body["categories"]] == ["맛", "간", "위생", "서비스", "기타"]
    assert categories["맛"]["count"] == 1
    assert categories["간"]["count"] == 1
    assert categories["위생"]["count"] == 1
    assert categories["서비스"]["count"] == 1
    assert categories["기타"]["count"] == 1
    assert categories["기타"]["comments"][0]["comment"] == "그냥 평범했어요"


def _corner_id(db_session, corner_name: str) -> int:
    """corner_master에서 직접 읽는다 — /api/analysis/corners는 배치 집계
    (daily_corner_stats)를 읽으므로 recompute 전에는 비어 있다."""
    from app.models.master import CornerMaster

    return db_session.query(CornerMaster).filter_by(corner_name=corner_name).one().corner_id


def _seed_headcount_trend_fixture(client):
    """코너 × 회사구분 교차 셀을 구분할 수 있게 시딩한다.

    MONDAY(평일): 한식×삼성전자(본사) 2건, 한식×삼성SDI(계열사) 1건,
                  분식×삼성전자(본사) 1건 — 중식
    MONDAY+5(토, 주말): 한식×삼성전자 1건 — 중식
    MONDAY(평일) 조식: 한식×삼성전자 1건
    """
    for i in range(2):
        _ingest_meal_log(client, f"HQ{i}", "맛남", company_name="삼성전자", corner_name="한식")
    _ingest_meal_log(client, "AF1", "맛남", company_name="삼성SDI", corner_name="한식")
    _ingest_meal_log(client, "HQ9", "맛남", company_name="삼성전자", corner_name="분식")
    _ingest_meal_log(
        client, "HQW", "맛남", company_name="삼성전자", corner_name="한식",
        eaten_date=MONDAY + dt.timedelta(days=5),
    )
    # 조식 1건 — meal_type 필터가 실제로 걸리는지 확인용
    client.post(
        "/api/ingest/meal-log",
        json={
            "rows": [
                {
                    "eaten_at": dt.datetime.combine(MONDAY, dt.time(8, 0)).isoformat(),
                    "employee_id": "HQB",
                    "meal_type": "조식",
                    "corner_name": "한식",
                    "taste_score": "맛남",
                    "company_name": "삼성전자",
                }
            ]
        },
        headers=AUTH_HEADERS,
    )


def _trend(client, **params):
    resp = client.get("/api/analysis/headcount-trend", params=params)
    assert resp.status_code == 200, resp.text
    return resp.json()


def test_headcount_trend_crosses_corner_and_division(client, db_session):
    """기존 엔드포인트로는 불가능했던 '코너 × 회사구분' 교차 필터(2026-08)."""
    _seed_headcount_trend_fixture(client)
    period = {"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=6)).isoformat()}

    # 한식 코너 + 본사만 = MONDAY 중식 2건 + MONDAY 조식 1건 + 토요일 1건 = 4건
    corner_id = _corner_id(db_session, "한식")
    rows = _trend(client, **period, corner_ids=[corner_id], divisions=["본사"])
    assert sum(r["headcount"] for r in rows) == 4

    # 같은 코너에서 계열사만 = 1건 (교차가 실제로 갈라지는지)
    rows = _trend(client, **period, corner_ids=[corner_id], divisions=["계열사"])
    assert sum(r["headcount"] for r in rows) == 1


def test_headcount_trend_meal_type_and_classification_filters(client):
    _seed_headcount_trend_fixture(client)
    period = {"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=6)).isoformat()}

    # 중식만 = 전체 6건 중 조식 1건 제외한 5건
    assert sum(r["headcount"] for r in _trend(client, **period, meal_types=["중식"])) == 5
    # 평일만 = 토요일 1건 제외한 5건
    assert sum(r["headcount"] for r in _trend(client, **period, classification="평일")) == 5
    # 주말+공휴일만 = 토요일 1건
    assert sum(r["headcount"] for r in _trend(client, **period, classification="주말+공휴일")) == 1


def test_headcount_trend_group_by_axes(client):
    _seed_headcount_trend_fixture(client)
    period = {"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=6)).isoformat()}

    by_division = _trend(client, **period, group_by="division")
    totals = {}
    for r in by_division:
        totals[r["series_label"]] = totals.get(r["series_label"], 0) + r["headcount"]
    assert totals == {"본사": 5, "계열사": 1}

    by_corner = _trend(client, **period, group_by="corner")
    corner_totals = {}
    for r in by_corner:
        corner_totals[r["series_label"]] = corner_totals.get(r["series_label"], 0) + r["headcount"]
    assert corner_totals == {"한식": 5, "분식": 1}

    # group_by=total이면 시리즈가 하나로 합쳐진다
    assert {r["series_label"] for r in _trend(client, **period)} == {"전체"}


def test_headcount_trend_total_matches_division_analysis(client):
    """신규 런타임 집계가 기존 배치 집계(daily_division_stats)와 합계가 일치하는지 —
    사번이 employee_master에 없을 때 Division.OTHER로 넣는 규칙까지 동일해야 한다."""
    _seed_headcount_trend_fixture(client)
    period_end = MONDAY + dt.timedelta(days=6)
    resp = client.post(
        "/api/analysis/daily-stats/recompute",
        params={"period_start": MONDAY.isoformat(), "period_end": period_end.isoformat()},
    )
    assert resp.status_code == 200, resp.text

    resp = client.get(
        "/api/analysis/divisions",
        params={"period_start": MONDAY.isoformat(), "period_end": period_end.isoformat()},
    )
    assert resp.status_code == 200
    legacy_total = sum(r["headcount"] for r in resp.json())

    new_total = sum(
        r["headcount"]
        for r in _trend(client, period_start=MONDAY.isoformat(), period_end=period_end.isoformat())
    )
    assert new_total == legacy_total


def test_weekly_congestion_forecast_skips_holidays_and_applies_multipliers(client, db_session):
    """현황 "금주 예상 식수" — 주 단위 래퍼(2026-08).

    휴일(토·일)은 식당이 안 열어 아예 빠지고, 날씨·연휴 전후 배수가 곱해진다.
    """
    from app.models.enums import HolidayType
    from app.models.master import HolidayCalendar

    # 과거 이력 시딩 — baseline이 0이 아니게 만든다
    for offset in (14, 7):
        for i in range(4):
            _ingest_meal_log(client, f"C{offset}_{i}", "맛남", eaten_date=MONDAY - dt.timedelta(days=offset))
    resp = client.post(
        "/api/analysis/daily-stats/recompute",
        params={
            "period_start": (MONDAY - dt.timedelta(days=14)).isoformat(),
            "period_end": (MONDAY - dt.timedelta(days=7)).isoformat(),
        },
    )
    assert resp.status_code == 200

    # MONDAY+7(월)을 공휴일로 등록 → 토·일·월 3일 연휴가 되어 직전 금요일이 "연휴 전"
    long_break_monday = MONDAY + dt.timedelta(days=7)
    db_session.add(
        HolidayCalendar(
            calendar_date=long_break_monday,
            holiday_type=HolidayType.STATUTORY,
            holiday_name="테스트공휴일",
        )
    )
    db_session.commit()

    resp = client.get(
        "/api/simulation/congestion-forecast/weekly",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": (MONDAY + dt.timedelta(days=6)).isoformat(),
            "meal_type": "중식",
            "weather": "비",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()

    returned_dates = {d["target_date"] for d in body["days"]}
    # 토(MONDAY+5)·일(MONDAY+6)은 휴일이라 빠진다
    assert (MONDAY + dt.timedelta(days=5)).isoformat() not in returned_dates
    assert (MONDAY + dt.timedelta(days=6)).isoformat() not in returned_dates
    assert MONDAY.isoformat() in returned_dates

    friday = next(d for d in body["days"] if d["target_date"] == (MONDAY + dt.timedelta(days=4)).isoformat())
    assert friday["holiday_adjacency"] == "연휴 전"
    # 비(0.90) × 연휴 전(0.85) = 0.765
    assert friday["applied_multiplier"] == 0.765

    monday = next(d for d in body["days"] if d["target_date"] == MONDAY.isoformat())
    assert monday["holiday_adjacency"] == "해당 없음"
    assert monday["applied_multiplier"] == 0.9  # 비만 적용
    assert monday["total_predicted_headcount"] > 0


def test_demoted_features_keep_working_apis(client, db_session):
    """2026-08 화면 재편: 아래 기능들은 **UI에서만 내리고 API·계산 로직은 유지**한다.

    "분석용으로 데이터는 가지고 있되 UI에 표현되는 것만 정리한다"는 결정이라,
    화면에서 안 보인다고 엔드포인트까지 같이 죽는 일이 없도록 여기서 고정한다.
    나중에 리포트나 다른 화면에서 되살릴 때 이 테스트가 계약서 역할을 한다.
    """
    _ingest_meal_log(client, "E1", "맛남", menu_name="제육볶음", corner_name="한식")
    corner_id = _corner_id(db_session, "한식")
    period = {"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=6)).isoformat()}

    # ① 코너 코어층(충성도) ② 서브속도 메뉴별
    assert client.get("/api/analysis/corners/core-layer-summary", params=period).status_code == 200
    assert client.get(f"/api/analysis/corners/{corner_id}/menu-throughput", params=period).status_code == 200

    # ④ 취향 군집 ⑤ 사번별 취향 벡터(개인정보 — UI만 내리고 API는 유지, 협의 결정)
    assert client.get("/api/analysis/users/taste-clusters").status_code == 200
    assert client.get("/api/analysis/users/E1/taste-profile").status_code in (200, 404)

    # ⑥ 캠퍼스 평균 음식벡터(레이더)
    assert client.get("/api/analysis/menus/food-vectors/average").status_code == 200

    # 시뮬레이션 탭은 없어지지만 what-if API는 남는다(현황 예측이 같은 로직을 쓴다)
    resp = client.post(
        "/api/simulation/what-if",
        json={"target_date": MONDAY.isoformat(), "meal_type": "중식", "weather": "맑음"},
    )
    assert resp.status_code == 200, resp.text


def test_voe_by_category_comment_includes_menu_name(client):
    # 어떤 메뉴에 대한 의견인지 알 수 있게 코멘트마다 menu_name도 함께 내려온다(2026-08).
    _ingest_meal_log(client, "E1", "맛남", comment="정말 맛있어요", menu_name="제육볶음")

    resp = client.get("/api/dashboard/voe-by-category", params={"period": f"{MONDAY.isoformat()[:7]}-01"})
    assert resp.status_code == 200
    body = resp.json()
    taste_comments = next(c for c in body["categories"] if c["category"] == "맛")["comments"]
    assert taste_comments[0]["menu_name"] == "제육볶음"


def test_voe_by_category_multi_label_comment_counted_in_both_categories(client):
    _ingest_meal_log(client, "E1", "개선", comment="맛없고 위생도 별로예요")

    resp = client.get("/api/dashboard/voe-by-category", params={"period": f"{MONDAY.isoformat()[:7]}-01"})
    body = resp.json()
    categories = {c["category"]: c for c in body["categories"]}

    assert body["total_comments"] == 1
    assert categories["맛"]["count"] == 1
    assert categories["위생"]["count"] == 1
    assert categories["간"]["count"] == 0


def test_voe_by_category_prefers_stored_llm_categories_over_rule_based(client, db_session):
    # "정말 좋아요"는 규칙 기반으로는 어느 카테고리에도 안 걸리지만(기타로 감),
    # voe_categories가 이미 채워져 있으면(LLM 배치 결과) 그걸 그대로 써야 한다.
    _ingest_meal_log(client, "E1", "맛남", comment="정말 좋아요")

    from app.models.logs import MealLog

    log = db_session.query(MealLog).filter_by(employee_id="E1").one()
    log.voe_categories = ["서비스"]
    db_session.commit()

    resp = client.get("/api/dashboard/voe-by-category", params={"period": f"{MONDAY.isoformat()[:7]}-01"})
    body = resp.json()
    categories = {c["category"]: c for c in body["categories"]}
    assert categories["서비스"]["count"] == 1
    assert categories["기타"]["count"] == 0


def test_voe_by_category_recompute_falls_back_to_rules_when_llm_unconfigured(client, db_session):
    _ingest_meal_log(client, "E1", "맛남", comment="정말 맛있어요")

    resp = client.post(
        "/api/dashboard/voe-by-category/recompute", params={"period": f"{MONDAY.isoformat()[:7]}-01"}
    )
    assert resp.status_code == 200
    assert resp.json()["classified_comments"] == 1

    from app.models.logs import MealLog

    log = db_session.query(MealLog).filter_by(employee_id="E1").one()
    assert log.voe_categories == ["맛"]


def test_voe_clusters_recompute_creates_clusters_from_comments(client, db_session):
    # 사내 LLM 미설정 환경이라 llm_client의 모의 임베딩/응답으로 배선만 검증한다
    # (voe_clustering.py의 실제 군집 품질은 사내 LLM 연동 후 별도 확인 필요).
    _ingest_meal_log(client, "E1", "맛남", comment="정말 맛있어요")
    _ingest_meal_log(client, "E2", "개선", comment="너무 짰어요")

    resp = client.post(
        "/api/dashboard/voe-clusters/recompute", params={"period": f"{MONDAY.isoformat()[:7]}-01"}
    )
    assert resp.status_code == 200
    assert resp.json()["clusters_created"] >= 1

    resp = client.get("/api/dashboard/voe-clusters", params={"period": f"{MONDAY.isoformat()[:7]}-01"})
    assert resp.status_code == 200
    clusters = resp.json()
    assert len(clusters) >= 1
    assert sum(c["comment_count"] for c in clusters) == 2


def test_voe_briefing_reports_no_clusters_before_clustering_runs(client):
    """§80: 클러스터링을 아직 안 돌렸으면 has_clusters=false로 화면이 구분할 수 있어야 한다."""
    resp = client.get("/api/dashboard/voe-briefing", params={"period": f"{MONDAY.isoformat()[:7]}-01"})
    assert resp.status_code == 200
    body = resp.json()
    assert body == {"has_clusters": False, "briefing": None, "briefing_computed_at": None}


def test_voe_briefing_recompute_reuses_existing_clusters(client):
    """§80: 재임베딩 없이 이미 계산된 MonthlyVoeCluster를 그대로 요약에 재사용한다."""
    _ingest_meal_log(client, "E1", "맛남", comment="정말 맛있어요")
    _ingest_meal_log(client, "E2", "개선", comment="너무 짰어요")
    period = f"{MONDAY.isoformat()[:7]}-01"

    resp = client.post("/api/dashboard/voe-clusters/recompute", params={"period": period})
    assert resp.status_code == 200
    assert resp.json()["clusters_created"] >= 1

    resp = client.post("/api/dashboard/voe-briefing/recompute", params={"period": period})
    assert resp.status_code == 200
    body = resp.json()
    assert body["has_clusters"] is True
    assert body["briefing"]

    resp = client.get("/api/dashboard/voe-briefing", params={"period": period})
    assert resp.status_code == 200
    body = resp.json()
    assert body["has_clusters"] is True
    assert body["briefing"]
    assert body["briefing_computed_at"] is not None


def test_average_menu_food_vector_uses_only_main_menus(client, db_session):
    _ingest_weekly_menu(client)  # 제육볶음(메인)/계란후라이(부찬), 한식, MONDAY

    from app.models.master import MenuMaster
    from app.services.food_vector import FOOD_VECTOR_DIM

    main = db_session.query(MenuMaster).filter_by(menu_name="제육볶음").one()
    side = db_session.query(MenuMaster).filter_by(menu_name="계란후라이").one()
    main.food_vector = [0.9] + [0.5] * (FOOD_VECTOR_DIM - 1)
    side.food_vector = [0.1] + [0.5] * (FOOD_VECTOR_DIM - 1)  # 부찬 — 평균에서 제외돼야 함
    db_session.commit()

    resp = client.get("/api/analysis/menus/food-vectors/average")
    assert resp.status_code == 200
    body = resp.json()
    assert body["sample_size"] == 1
    assert body["average"][0] == pytest.approx(0.9, abs=1e-4)
    assert body["dimensions"] == list(body["labels_ko"].keys())
    assert body["bias_description"] is not None


def test_average_menu_food_vector_empty_when_no_tagged_main_menus(client):
    resp = client.get("/api/analysis/menus/food-vectors/average")
    assert resp.status_code == 200
    body = resp.json()
    assert body["sample_size"] == 0
    assert body["bias_description"] is None


def test_menu_comments_returns_recent_comments_newest_first(client):
    _ingest_meal_log(client, "E1", "맛남", comment="맛있어요", eaten_date=MONDAY, menu_name="제육볶음")
    _ingest_meal_log(
        client, "E2", "개선", comment="좀 짰어요", eaten_date=MONDAY + dt.timedelta(days=1), menu_name="제육볶음"
    )
    _ingest_meal_log(client, "E3", "맛남", comment=None, eaten_date=MONDAY, menu_name="제육볶음")

    resp = client.get("/api/dashboard/menu-comments/제육볶음")
    assert resp.status_code == 200
    body = resp.json()
    assert [c["comment"] for c in body] == ["좀 짰어요", "맛있어요"]
    assert body[0]["taste_score"] == "개선"


def test_menu_comments_unknown_menu_returns_404(client):
    resp = client.get("/api/dashboard/menu-comments/없는메뉴")
    assert resp.status_code == 404


def test_corner_main_menu_by_date_returns_main_menu_only(client):
    resp = client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                {
                    "plan_date": MONDAY.isoformat(),
                    "meal_type": "중식",
                    "corner_name": "한식",
                    "menu_name": "제육볶음",
                    "menu_role": "메인",
                    "source_row_raw": "제육볶음",
                },
                {
                    "plan_date": MONDAY.isoformat(),
                    "meal_type": "중식",
                    "corner_name": "한식",
                    "menu_name": "김치",
                    "menu_role": "부찬",
                    "source_row_raw": "김치",
                },
            ]
        },
        headers=AUTH_HEADERS,
    )
    assert resp.status_code == 200

    resp = client.get(
        "/api/analysis/corners/main-menu-by-date",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert len(body) == 1
    assert body[0]["menu_name"] == "제육볶음"
    assert body[0]["plan_date"] == MONDAY.isoformat()


def test_top_menus_by_headcount_ranks_by_appearance_count(client):
    _ingest_meal_log(client, "E1", "맛남", eaten_date=MONDAY, menu_name="제육볶음")
    _ingest_meal_log(client, "E2", "맛남", eaten_date=MONDAY, menu_name="제육볶음")
    _ingest_meal_log(client, "E3", "맛남", eaten_date=MONDAY, menu_name="돈까스")

    resp = client.get(
        "/api/analysis/menus/top-by-headcount",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "top_n": 5},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body[0]["menu_name"] == "제육볶음"
    assert body[0]["headcount"] == 2
    assert body[1]["menu_name"] == "돈까스"
    assert body[1]["headcount"] == 1


def test_top_menus_by_headcount_respects_top_n(client):
    _ingest_meal_log(client, "E1", "맛남", eaten_date=MONDAY, menu_name="제육볶음")
    _ingest_meal_log(client, "E2", "맛남", eaten_date=MONDAY, menu_name="돈까스")

    resp = client.get(
        "/api/analysis/menus/top-by-headcount",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "top_n": 1},
    )
    assert resp.status_code == 200
    assert len(resp.json()) == 1


def test_top_menus_by_headcount_outside_period_returns_empty(client):
    _ingest_meal_log(client, "E1", "맛남", eaten_date=MONDAY, menu_name="제육볶음")

    resp = client.get(
        "/api/analysis/menus/top-by-headcount",
        params={
            "period_start": (MONDAY - dt.timedelta(days=30)).isoformat(),
            "period_end": (MONDAY - dt.timedelta(days=1)).isoformat(),
        },
    )
    assert resp.status_code == 200
    assert resp.json() == []


def test_congestion_forecast_uses_family_day_specific_baseline(client, db_session):
    from app.models.enums import MealType
    from app.models.master import CornerMaster
    from app.models.stats import DailyCornerStats
    from app.services.holidays import family_day_of_month

    resp = client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                {
                    "plan_date": MONDAY.isoformat(),
                    "meal_type": "중식",
                    "corner_name": "한식",
                    "menu_name": "제육볶음",
                    "menu_role": "메인",
                    "source_row_raw": "제육볶음",
                }
            ]
        },
        headers=AUTH_HEADERS,
    )
    assert resp.status_code == 200
    corner = db_session.query(CornerMaster).filter_by(corner_name="한식").one()

    # 평일 이력: 낮은 식수(10명) 8회
    for i in range(1, 9):
        db_session.add(
            DailyCornerStats(
                stat_date=MONDAY - dt.timedelta(days=i),
                corner_id=corner.corner_id,
                meal_type=MealType.LUNCH,
                headcount=10,
                is_holiday=False,
            )
        )
    # 패밀리데이 이력: 과거 8개월치, 뚜렷이 다른 식수(200명)
    year, month = MONDAY.year, MONDAY.month
    for _ in range(8):
        month -= 1
        if month == 0:
            month, year = 12, year - 1
        db_session.add(
            DailyCornerStats(
                stat_date=family_day_of_month(year, month),
                corner_id=corner.corner_id,
                meal_type=MealType.LUNCH,
                headcount=200,
                is_holiday=False,
            )
        )
    db_session.commit()

    target_family_day = family_day_of_month(MONDAY.year, MONDAY.month)  # 2026-07-24
    resp = client.get(
        "/api/simulation/congestion-forecast",
        params={"target_date": target_family_day.isoformat(), "meal_type": "중식"},
    )
    assert resp.status_code == 200
    corner_row = next(c for c in resp.json()["corners"] if c["corner_name"] == "한식")
    # 패밀리데이 이력(200명)만 baseline에 반영되고 평일 이력(10명)과 안 섞여야 한다
    assert corner_row["predicted_headcount"] == pytest.approx(200.0, abs=0.5)


def test_what_if_reports_family_day_classification(client):
    from app.services.holidays import family_day_of_month

    target_family_day = family_day_of_month(MONDAY.year, MONDAY.month)
    resp = client.post(
        "/api/simulation/what-if",
        json={"target_date": target_family_day.isoformat(), "meal_type": "중식", "weather": "맑음"},
    )
    assert resp.status_code == 200
    assert resp.json()["classification"] == "패밀리데이"


def _ingest_meal_log_with_meal_type(client, employee_id, taste, meal_type, menu_name, eaten_date=MONDAY):
    rows = [
        {
            "eaten_at": dt.datetime.combine(eaten_date, dt.time(8, 0, 0)).isoformat(),
            "employee_id": employee_id,
            "meal_type": meal_type,
            "corner_name": "한식",
            "taste_score": taste,
            "menu_name": menu_name,
        }
    ]
    resp = client.post("/api/ingest/meal-log", json={"rows": rows}, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text


def test_weekly_summary_filters_by_meal_types(client, db_session):
    from app.services.aggregation import aggregate_daily_stats

    _ingest_meal_log_with_meal_type(client, "E1", "맛남", "조식", "토스트")
    _ingest_meal_log_with_meal_type(client, "E2", "맛남", "중식", "제육볶음")
    aggregate_daily_stats(db_session, MONDAY)

    breakfast_only = client.get(
        "/api/dashboard/weekly-summary",
        params={"start_date": MONDAY.isoformat(), "end_date": MONDAY.isoformat(), "meal_types": ["조식"]},
    )
    assert breakfast_only.status_code == 200
    assert breakfast_only.json()[0]["headcount"] == 1

    combined = client.get(
        "/api/dashboard/weekly-summary",
        params={
            "start_date": MONDAY.isoformat(),
            "end_date": MONDAY.isoformat(),
            "meal_types": ["조식", "중식"],
        },
    )
    assert combined.json()[0]["headcount"] == 2

    unfiltered = client.get(
        "/api/dashboard/weekly-summary",
        params={"start_date": MONDAY.isoformat(), "end_date": MONDAY.isoformat()},
    )
    assert unfiltered.json()[0]["headcount"] == 2


def test_corner_analysis_trend_filters_by_meal_types(client, db_session):
    from app.services.aggregation import aggregate_daily_stats

    _ingest_meal_log_with_meal_type(client, "E1", "맛남", "조식", "토스트")
    _ingest_meal_log_with_meal_type(client, "E2", "맛남", "중식", "제육볶음")
    aggregate_daily_stats(db_session, MONDAY)

    resp = client.get(
        "/api/analysis/corners/trend",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": MONDAY.isoformat(),
            "granularity": "daily",
            "meal_types": ["조식"],
        },
    )
    assert resp.status_code == 200
    hansik = next(r for r in resp.json() if r["corner_name"] == "한식")
    assert hansik["headcount"] == 1


def test_menu_performance_by_meal_type_filters_to_selected_meal(client):
    _ingest_meal_log_with_meal_type(client, "E1", "맛남", "조식", "토스트")
    _ingest_meal_log_with_meal_type(client, "E2", "맛남", "조식", "토스트")
    _ingest_meal_log_with_meal_type(client, "E3", "맛남", "중식", "제육볶음")

    resp = client.get(
        "/api/analysis/menu-performance/by-meal-type",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "meal_type": "조식"},
    )
    assert resp.status_code == 200
    body = resp.json()
    menu_names = {r["menu_name"] for r in body}
    assert menu_names == {"토스트"}
    row = next(r for r in body if r["menu_name"] == "토스트")
    assert row["total_headcount"] == 2


def test_menu_performance_by_meal_type_empty_when_no_logs_for_meal(client):
    _ingest_meal_log_with_meal_type(client, "E1", "맛남", "중식", "제육볶음")

    resp = client.get(
        "/api/analysis/menu-performance/by-meal-type",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "meal_type": "석식"},
    )
    assert resp.status_code == 200
    assert resp.json() == []


def test_menu_performance_by_meal_type_computes_quadrant_within_meal_type(client):
    for i in range(6):
        _ingest_meal_log_with_meal_type(client, f"A{i}", "맛남", "조식", "인기메뉴A")
    for i in range(2):
        _ingest_meal_log_with_meal_type(client, f"B{i}", "개선", "조식", "비인기메뉴B")

    resp = client.get(
        "/api/analysis/menu-performance/by-meal-type",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "meal_type": "조식"},
    )
    assert resp.status_code == 200
    body = resp.json()
    row_a = next(r for r in body if r["menu_name"] == "인기메뉴A")
    row_b = next(r for r in body if r["menu_name"] == "비인기메뉴B")
    assert row_a["total_headcount"] == 6
    assert row_b["total_headcount"] == 2
    assert row_a["quadrant"] is not None
    assert row_b["quadrant"] is not None


# ---------------------------------------------------------------------------
# 2순위: 메뉴 회전 이력 + 건강가든 텍스트 입력 (2026-08)
# ---------------------------------------------------------------------------


def _plan_row(plan_date, menu_name, menu_role, corner_name="한식", meal_type="중식"):
    return {
        "plan_date": plan_date.isoformat(),
        "meal_type": meal_type,
        "corner_name": corner_name,
        "menu_name": menu_name,
        "menu_role": menu_role,
        "source_row_raw": menu_name,
    }


def _rotation(client, **params):
    resp = client.get("/api/analysis/weekly-menu/rotation", params=params)
    assert resp.status_code == 200, resp.text
    return resp.json()


def test_rotation_flags_menu_replanned_too_soon(client):
    """직전 편성 이후 14일을 못 채우면 "재편성 과다"."""
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY - dt.timedelta(days=5), "돈까스", "메인"),
                _plan_row(MONDAY, "돈까스", "메인"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    data = _rotation(client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat())
    items = [i for i in data["items"] if i["menu_name"] == "돈까스"]
    assert len(items) == 1
    assert items[0]["flag"] == "재편성 과다"
    assert items[0]["gap_days"] == 5
    assert items[0]["previous_date"] == (MONDAY - dt.timedelta(days=5)).isoformat()


def test_rotation_includes_avg_satisfaction_and_headcount(client):
    """§81: 메뉴 중복점검 재설계 — Top5/기준 미달 목록에 쓸 만족도·식수 필드."""
    client.post(
        "/api/ingest/weekly-menu",
        json={"rows": [_plan_row(MONDAY, "돈까스", "메인")]},
        headers=AUTH_HEADERS,
    )
    _ingest_meal_log(client, "R1", "맛남", eaten_date=MONDAY, menu_name="돈까스")
    _ingest_meal_log(client, "R2", "보통", eaten_date=MONDAY, menu_name="돈까스")

    data = _rotation(client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat())
    item = next(i for i in data["items"] if i["menu_name"] == "돈까스")
    assert item["recent_avg_headcount"] == 2.0
    assert item["avg_satisfaction"] == 4.0  # 맛남=5, 보통=3 평균


def test_rotation_headcount_and_satisfaction_null_without_meal_log(client):
    """취식 기록이 아예 없는 메뉴는 만족도·식수 둘 다 null이어야 한다(0이 아님)."""
    client.post(
        "/api/ingest/weekly-menu",
        json={"rows": [_plan_row(MONDAY, "잡채", "메인")]},
        headers=AUTH_HEADERS,
    )
    data = _rotation(client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat())
    item = next(i for i in data["items"] if i["menu_name"] == "잡채")
    assert item["recent_avg_headcount"] is None
    assert item["avg_satisfaction"] is None


def test_rotation_history_outside_period_is_not_returned_as_item(client):
    """과거 편성은 판정 기준으로만 쓰고 결과 목록엔 안 나와야 한다."""
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY - dt.timedelta(days=60), "갈비탕", "메인"),
                _plan_row(MONDAY, "갈비탕", "메인"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    data = _rotation(client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat())
    dates = {i["plan_date"] for i in data["items"]}
    assert dates == {MONDAY.isoformat()}
    galbi = next(i for i in data["items"] if i["menu_name"] == "갈비탕")
    assert galbi["flag"] == "적정"  # 60일 만이면 절대 기준 통과, 평균 주기는 미산출
    assert galbi["gap_days"] == 60


def test_same_menu_in_two_corners_on_one_day_is_not_a_duplicate(client):
    """다른 코너에 같은 메뉴가 깔린 건 중복이 아니다.

    ⚠️ 이 테스트는 예전에 정반대를 주장했다 — "같은 날 다른 코너면 같은 날 중복"이
    의도된 동작이었다. 2026-08 담당자 기준이 바뀌었다: "포기김치가 다른 코너에서
    각각 나왔다고 중복이면 안 되고". 코너는 서로 다른 선택지지 중복이 아니다.
    """
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY, "김치찌개", "메인", corner_name="한식"),
                _plan_row(MONDAY, "김치찌개", "메인", corner_name="일품"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    data = _rotation(client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat())
    kimchi = [i for i in data["items"] if i["menu_name"] == "김치찌개"]
    assert len(kimchi) == 2
    assert all(i["flag"] != "같은 날 중복" for i in kimchi), "코너가 다른데 중복으로 잡혔다"


def test_side_dish_clashing_with_health_garden_same_day_is_a_duplicate(client):
    """건강가든은 누구나 가져가는 공용이라 코너를 가로질러 겹침으로 본다.

    담당자: "건강가든하고만 중복 봐야함".
    """
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY, "제육볶음", "메인", corner_name="한식"),
                _plan_row(MONDAY, "시금치나물", "부찬", corner_name="한식"),
                _plan_row(MONDAY, "시금치나물", "건강가든", corner_name="그린미트"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    data = _rotation(client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat())
    spinach = [i for i in data["items"] if i["menu_name"] == "시금치나물"]
    assert spinach, "시금치나물이 회전 이력에 없다"
    assert any(i["flag"] == "같은 날 중복" for i in spinach)


def test_rotation_reports_overused_menus_in_period(client):
    """한 주에 4번 들어간 부찬은 overused로 잡힌다(기본 임계 3회 초과)."""
    rows = [
        _plan_row(MONDAY + dt.timedelta(days=i), "시금치나물", "부찬") for i in range(4)
    ]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    data = _rotation(
        client,
        period_start=MONDAY.isoformat(),
        period_end=(MONDAY + dt.timedelta(days=5)).isoformat(),
    )
    overused = {o["menu_name"]: o for o in data["overused"]}
    assert "시금치나물" in overused
    assert overused["시금치나물"]["count"] == 4


# ---------------------------------------------------------------------------
# 자주 반복되는 부찬 랭킹 — 임의 기간 (2026-08)
# ---------------------------------------------------------------------------
# 담당자: "부찬 중복 볼 때 보기가 너무 불편함, 정말 자주 나오고 돌려막기한 부찬을
# 보고싶어". /weekly-menu/rotation의 overused는 화면이 요청한 한 주치만 보여줘서
# "지난 3개월 동안 자주 반복됐다"는 그림이 안 나왔다. 이 엔드포인트는 담당자가
# 고른 임의 기간 하나로 그 랭킹을 낸다 — find_overused_menus 자체(코너 스코프·
# 고유 날짜·정렬)는 이미 검증돼 있으므로(tests/test_menu_rotation.py), 여기선
# 배선(역할 필터·코너 필터·정렬 전달)만 확인한다.


def _repeated(client, **params):
    resp = client.get("/api/analysis/weekly-menu/repeated-side-dishes", params=params)
    assert resp.status_code == 200, resp.text
    return resp.json()


def test_repeated_side_dishes_sorted_by_count_desc(client):
    rows = [
        _plan_row(MONDAY + dt.timedelta(days=i), "시금치나물", "부찬") for i in range(5)
    ] + [_plan_row(MONDAY + dt.timedelta(days=i), "콩나물무침", "부찬") for i in range(2)]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)

    data = _repeated(
        client, period_start=MONDAY.isoformat(), period_end=(MONDAY + dt.timedelta(days=6)).isoformat()
    )
    names = [i["menu_name"] for i in data["items"]]
    assert names.index("시금치나물") < names.index("콩나물무침")
    counts = {i["menu_name"]: i["count"] for i in data["items"]}
    assert counts["시금치나물"] == 5
    assert counts["콩나물무침"] == 2


def test_repeated_side_dishes_excludes_main_menu(client):
    """부찬 랭킹이라 메인은 아무리 자주 나와도 빠져야 한다."""
    rows = [_plan_row(MONDAY + dt.timedelta(days=i), "돈까스", "메인") for i in range(5)]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)

    data = _repeated(
        client, period_start=MONDAY.isoformat(), period_end=(MONDAY + dt.timedelta(days=6)).isoformat()
    )
    assert all(i["menu_name"] != "돈까스" for i in data["items"])


def test_repeated_side_dishes_corner_filter_narrows_results(client, db_session):
    """포기김치가 다른 코너에서 각각 나온 건 그 코너 안에서만 세야 한다(§132)."""
    rows = [
        _plan_row(MONDAY + dt.timedelta(days=i), "포기김치", "부찬", corner_name="한식") for i in range(3)
    ] + [
        _plan_row(MONDAY + dt.timedelta(days=i), "포기김치", "부찬", corner_name="일품") for i in range(2)
    ]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    hansik_id = _corner_id(db_session, "한식")

    data = _repeated(
        client,
        period_start=MONDAY.isoformat(),
        period_end=(MONDAY + dt.timedelta(days=6)).isoformat(),
        corner_id=hansik_id,
    )
    assert len(data["items"]) == 1
    assert data["items"][0]["corner_name"] == "한식"
    assert data["items"][0]["count"] == 3


def test_repeated_side_dishes_includes_health_garden(client):
    rows = [
        _plan_row(MONDAY + dt.timedelta(days=i), "대추차", "건강가든") for i in range(3)
    ]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)

    data = _repeated(
        client, period_start=MONDAY.isoformat(), period_end=(MONDAY + dt.timedelta(days=6)).isoformat()
    )
    item = next(i for i in data["items"] if i["menu_name"] == "대추차")
    assert item["menu_role"] == "건강가든"
    assert item["count"] == 3


def test_repeated_side_dishes_counts_unique_dates_not_rows(client):
    """같은 날 두 끼니(조식/중식)에 겹쳐 들어가도 그날은 1번으로 센다."""
    rows = [
        _plan_row(MONDAY, "김치", "부찬", meal_type="조식"),
        _plan_row(MONDAY, "김치", "부찬", meal_type="중식"),
        _plan_row(MONDAY + dt.timedelta(days=1), "김치", "부찬"),
    ]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)

    data = _repeated(
        client, period_start=MONDAY.isoformat(), period_end=(MONDAY + dt.timedelta(days=1)).isoformat()
    )
    item = next(i for i in data["items"] if i["menu_name"] == "김치")
    assert item["count"] == 2


# ---------------------------------------------------------------------------
# §80: 부찬 클릭 상세(날짜/코너/메인메뉴) + 연결 메인 만족도
# ---------------------------------------------------------------------------


def _side_dish_detail(client, **params):
    resp = client.get("/api/analysis/weekly-menu/side-dish-detail", params=params)
    assert resp.status_code == 200, resp.text
    return resp.json()


def test_side_dish_detail_returns_date_corner_main_menu_and_satisfaction(client):
    """단무지 클릭 → 그 날짜/코너/메인메뉴 목록이 나오고, 그날 메인의 평균
    만족도도 같이 붙어야 한다."""
    rows = [
        _plan_row(MONDAY, "짜장면", "메인"),
        _plan_row(MONDAY, "단무지", "부찬"),
    ]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    _ingest_meal_log(client, "E1", "맛남", menu_name="짜장면", corner_name="한식", eaten_date=MONDAY)
    _ingest_meal_log(client, "E2", "보통", menu_name="짜장면", corner_name="한식", eaten_date=MONDAY)

    data = _side_dish_detail(
        client,
        menu_name="단무지",
        corner_name="한식",
        period_start=MONDAY.isoformat(),
        period_end=(MONDAY + dt.timedelta(days=6)).isoformat(),
    )
    assert len(data["pairings"]) == 1
    pairing = data["pairings"][0]
    assert pairing["plan_date"] == MONDAY.isoformat()
    assert pairing["corner_name"] == "한식"
    assert pairing["main_menu_name"] == "짜장면"
    assert pairing["main_avg_satisfaction"] == 4.0  # (5+3)/2


def test_side_dish_detail_health_garden_matches_every_corner_same_day(client):
    """건강가든은 코너 무관 공용이라(§132) 그날 모든 코너의 메인과 매칭돼야 한다."""
    rows = [
        _plan_row(MONDAY, "제육볶음", "메인", corner_name="한식"),
        _plan_row(MONDAY, "생선구이", "메인", corner_name="일식"),
        _plan_row(MONDAY, "대추차", "건강가든", corner_name="한식"),
    ]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)

    data = _side_dish_detail(
        client,
        menu_name="대추차",
        corner_name="한식",
        period_start=MONDAY.isoformat(),
        period_end=(MONDAY + dt.timedelta(days=6)).isoformat(),
    )
    main_menu_names = {p["main_menu_name"] for p in data["pairings"]}
    assert main_menu_names == {"제육볶음", "생선구이"}


def test_repeated_side_dishes_includes_avg_main_satisfaction(client):
    rows = [_plan_row(MONDAY, "짜장면", "메인")] + [
        _plan_row(MONDAY + dt.timedelta(days=i), "단무지", "부찬") for i in range(3)
    ]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    _ingest_meal_log(client, "E1", "맛남", menu_name="짜장면", corner_name="한식", eaten_date=MONDAY)

    data = _repeated(
        client, period_start=MONDAY.isoformat(), period_end=(MONDAY + dt.timedelta(days=6)).isoformat()
    )
    item = next(i for i in data["items"] if i["menu_name"] == "단무지")
    assert item["avg_main_satisfaction"] == 5.0


def test_health_garden_text_input_replaces_slot_and_feeds_rotation(client, db_session):
    """건강가든 텍스트 입력 → 식단표 조회에 반영되고 회전 판정에도 들어간다."""
    _ingest_weekly_menu(client)
    corner_id = _corner_id(db_session, "한식")

    resp = client.put(
        "/api/analysis/weekly-menu/health-garden",
        json={
            "plan_date": MONDAY.isoformat(),
            "corner_id": corner_id,
            "meal_type": "중식",
            "menu_names_raw": "구운채소, 두부샐러드\n닭가슴살",
        },
    )
    assert resp.status_code == 200, resp.text
    assert [i["menu_name"] for i in resp.json()["items"]] == ["구운채소", "두부샐러드", "닭가슴살"]

    listed = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    ).json()
    slot = next(s for s in listed if s["corner_id"] == corner_id)
    assert [i["menu_name"] for i in slot["health_garden"]] == ["구운채소", "두부샐러드", "닭가슴살"]
    # 부찬과 섞이지 않아야 부찬 조합 비교가 오염되지 않는다
    assert [i["menu_name"] for i in slot["sides"]] == ["계란후라이"]

    # 건강가든도 회전 판정 대상이다("메인/부찬/건강가든 조합 중복 최소화")
    rotation = _rotation(client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat())
    garden = [i for i in rotation["items"] if i["menu_role"] == "건강가든"]
    assert {i["menu_name"] for i in garden} == {"구운채소", "두부샐러드", "닭가슴살"}

    # 전체 교체 — 다시 보내면 이전 목록은 사라진다
    client.put(
        "/api/analysis/weekly-menu/health-garden",
        json={
            "plan_date": MONDAY.isoformat(),
            "corner_id": corner_id,
            "meal_type": "중식",
            "menu_names_raw": "구운채소",
        },
    )
    listed = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    ).json()
    slot = next(s for s in listed if s["corner_id"] == corner_id)
    assert [i["menu_name"] for i in slot["health_garden"]] == ["구운채소"]


def test_health_garden_empty_input_clears_the_slot(client, db_session):
    _ingest_weekly_menu(client)
    corner_id = _corner_id(db_session, "한식")
    body = {
        "plan_date": MONDAY.isoformat(),
        "corner_id": corner_id,
        "meal_type": "중식",
        "menu_names_raw": "구운채소",
    }
    client.put("/api/analysis/weekly-menu/health-garden", json=body)
    client.put("/api/analysis/weekly-menu/health-garden", json={**body, "menu_names_raw": "  "})

    listed = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    ).json()
    slot = next(s for s in listed if s["corner_id"] == corner_id)
    assert slot["health_garden"] == []


def test_menu_combinations_corner_filter_narrows_slots(client, db_session):
    """같은 메인이 두 코너에서 다른 부찬과 나오면 코너 필터로 분리해 볼 수 있다."""
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY, "제육볶음", "메인", corner_name="한식"),
                _plan_row(MONDAY, "계란후라이", "부찬", corner_name="한식"),
                _plan_row(MONDAY + dt.timedelta(days=1), "제육볶음", "메인", corner_name="일품"),
                _plan_row(MONDAY + dt.timedelta(days=1), "미역국", "부찬", corner_name="일품"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    params = {
        "period_start": MONDAY.isoformat(),
        "period_end": (MONDAY + dt.timedelta(days=5)).isoformat(),
    }
    everything = client.get("/api/analysis/menu-combinations/제육볶음", params=params).json()
    assert len(everything["combos"]) == 2

    only_hansik = client.get(
        "/api/analysis/menu-combinations/제육볶음",
        params={**params, "corner_id": _corner_id(db_session, "한식")},
    ).json()
    assert only_hansik["corner_id"] is not None
    assert [c["sides"] for c in only_hansik["combos"]] == [["계란후라이"]]


def _combination_check(client, **params):
    resp = client.get("/api/analysis/weekly-menu/combination-check", params=params)
    assert resp.status_code == 200, resp.text
    return resp.json()


def test_combination_check_flags_ingredient_overlap_between_main_and_side(client):
    """담당자가 든 실제 예시 — 콩나물국밥(메인) + 콩나물무침(부찬)."""
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY, "콩나물국밥", "메인"),
                _plan_row(MONDAY, "콩나물무침", "부찬"),
                _plan_row(MONDAY, "깍두기", "부찬"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    data = _combination_check(
        client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat()
    )
    slot = data["slots"][0]
    shared = [c["shared"] for c in slot["ingredient_clashes"]]
    assert ["콩나물"] in shared
    assert slot["main"] == "콩나물국밥"


def test_combination_check_includes_health_garden_in_the_comparison(client, db_session):
    """건강가든도 부찬과 같이 본다 — 요청이 "메인/부찬/건강가든 조합"이었다."""
    client.post(
        "/api/ingest/weekly-menu",
        json={"rows": [_plan_row(MONDAY, "두부김치", "메인")]},
        headers=AUTH_HEADERS,
    )
    corner_id = _corner_id(db_session, "한식")
    client.put(
        "/api/analysis/weekly-menu/health-garden",
        json={
            "plan_date": MONDAY.isoformat(),
            "corner_id": corner_id,
            "meal_type": "중식",
            "menu_names_raw": "두부샐러드",
        },
    )
    data = _combination_check(
        client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat()
    )
    slot = data["slots"][0]
    assert slot["health_garden"] == ["두부샐러드"]
    pairs = {(c["menu_a"], c["menu_b"]) for c in slot["ingredient_clashes"]}
    assert ("두부김치", "두부샐러드") in pairs


def test_combination_check_reports_untagged_menus_instead_of_passing_them(client):
    """food_vector 미태깅 메뉴를 조용히 '충돌 없음'으로 넘기면 안 된다."""
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                # 규칙 키워드에 안 걸리는 이름 → food_vector가 NULL로 남는다
                _plan_row(MONDAY, "그라탱", "메인"),
                _plan_row(MONDAY, "라따뚜이", "부찬"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    data = _combination_check(
        client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat()
    )
    assert data["untagged_menu_count"] == 2
    assert set(data["slots"][0]["untagged"]) == {"그라탱", "라따뚜이"}


def test_combination_check_sorts_slots_with_more_clashes_first(client):
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                # 월: 충돌 없음
                _plan_row(MONDAY, "돈까스", "메인", corner_name="일품"),
                _plan_row(MONDAY, "미역국", "부찬", corner_name="일품"),
                # 화: 재료 중복 있음
                _plan_row(MONDAY + dt.timedelta(days=1), "감자탕", "메인", corner_name="일품"),
                _plan_row(MONDAY + dt.timedelta(days=1), "감자조림", "부찬", corner_name="일품"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    data = _combination_check(
        client,
        period_start=MONDAY.isoformat(),
        period_end=(MONDAY + dt.timedelta(days=5)).isoformat(),
    )
    assert len(data["slots"][0]["ingredient_clashes"]) > 0


def test_bulk_combo_loader_matches_single_menu_loader(client, db_session):
    """랭킹(bulk)과 상세(단건)가 같은 ComboDay를 내야 한다.

    두 경로가 갈라지면 화면에서 랭킹과 상세가 서로 다른 만족도를 보여준다.
    bulk는 쿼리 3개로 끝내는 최적화 버전이라 이 동치성이 유일한 안전장치다.
    """
    from app.models.master import MenuMaster
    from app.services.menu_combination import (
        build_side_combos_bulk,
        build_side_combos_for_main_menu,
    )

    rows = [
        _plan_row(MONDAY, "제육볶음", "메인"),
        _plan_row(MONDAY, "계란후라이", "부찬"),
        _plan_row(MONDAY + dt.timedelta(days=2), "제육볶음", "메인"),
        _plan_row(MONDAY + dt.timedelta(days=2), "미역국", "부찬"),
        _plan_row(MONDAY + dt.timedelta(days=1), "돈까스", "메인", corner_name="일품"),
        _plan_row(MONDAY + dt.timedelta(days=1), "단무지", "부찬", corner_name="일품"),
    ]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    # 취식 + 맛평가를 섞어 넣어 평균 만족도가 실제로 계산되게 한다
    _ingest_meal_log(client, "E1", "맛남", corner_name="한식", menu_name="제육볶음")
    _ingest_meal_log(client, "E2", "보통", corner_name="한식", menu_name="제육볶음")
    _ingest_meal_log(
        client, "E3", "개선", corner_name="한식", menu_name="제육볶음",
        eaten_date=MONDAY + dt.timedelta(days=2),
    )

    period_start, period_end = MONDAY, MONDAY + dt.timedelta(days=5)
    bulk = build_side_combos_bulk(db_session, period_start, period_end)

    menus = {m.menu_name: m.menu_id for m in db_session.query(MenuMaster).all()}
    for menu_name in ("제육볶음", "돈까스"):
        menu_id = menus[menu_name]
        single = build_side_combos_for_main_menu(db_session, menu_id, period_start, period_end)
        assert sorted(bulk.get(menu_id, []), key=lambda d: d.plan_date) == sorted(
            single, key=lambda d: d.plan_date
        ), f"{menu_name}에서 bulk와 단건 결과가 다름"


def test_bulk_combo_loader_respects_corner_filter(client, db_session):
    from app.models.master import MenuMaster
    from app.services.menu_combination import build_side_combos_bulk

    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY, "제육볶음", "메인", corner_name="한식"),
                _plan_row(MONDAY + dt.timedelta(days=1), "제육볶음", "메인", corner_name="일품"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    menu_id = db_session.query(MenuMaster).filter_by(menu_name="제육볶음").one().menu_id
    everything = build_side_combos_bulk(db_session, MONDAY, MONDAY + dt.timedelta(days=5))
    only_hansik = build_side_combos_bulk(
        db_session, MONDAY, MONDAY + dt.timedelta(days=5), corner_id=_corner_id(db_session, "한식")
    )
    assert len(everything[menu_id]) == 2
    assert len(only_hansik[menu_id]) == 1


def test_spread_ranking_puts_biggest_satisfaction_gap_first(client):
    """부찬 조합에 따라 만족도가 크게 갈리는 메인메뉴가 맨 위에 와야 한다."""
    rows = [
        # 제육볶음: 조합A는 맛있음, 조합B는 맛없음 → 편차 큼
        _plan_row(MONDAY, "제육볶음", "메인"),
        _plan_row(MONDAY, "계란후라이", "부찬"),
        _plan_row(MONDAY + dt.timedelta(days=1), "제육볶음", "메인"),
        _plan_row(MONDAY + dt.timedelta(days=1), "단무지", "부찬"),
    ]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    _ingest_meal_log(client, "E1", "맛남", corner_name="한식", menu_name="제육볶음")
    _ingest_meal_log(
        client, "E2", "개선", corner_name="한식", menu_name="제육볶음",
        eaten_date=MONDAY + dt.timedelta(days=1),
    )

    resp = client.get(
        "/api/analysis/menu-combinations/spread-ranking",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": (MONDAY + dt.timedelta(days=5)).isoformat(),
            "min_day_count": 1,
        },
    )
    assert resp.status_code == 200, resp.text
    items = resp.json()["items"]
    assert items, "편차가 있는 메뉴가 하나도 안 나왔다"
    top = items[0]
    assert top["menu_name"] == "제육볶음"
    assert top["spread"] > 0
    # best/worst가 실제로 서로 다른 조합이어야 한다
    assert top["best"]["sides"] != top["worst"]["sides"]
    assert top["best"]["avg_satisfaction"] > top["worst"]["avg_satisfaction"]


def test_spread_ranking_skips_menus_with_single_scored_combo(client):
    """평가 있는 조합이 1개뿐이면 비교가 불가능하므로 랭킹에서 빠진다."""
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY, "갈비탕", "메인"),
                _plan_row(MONDAY, "깍두기", "부찬"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    _ingest_meal_log(client, "E1", "맛남", corner_name="한식", menu_name="갈비탕")
    resp = client.get(
        "/api/analysis/menu-combinations/spread-ranking",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": (MONDAY + dt.timedelta(days=5)).isoformat(),
            "min_day_count": 1,
        },
    )
    assert [i["menu_name"] for i in resp.json()["items"]] == []


def _plan_performance(client, **params):
    resp = client.get("/api/analysis/menu-plan/performance", params=params)
    assert resp.status_code == 200, resp.text
    return resp.json()


def test_plan_performance_shows_menus_that_were_planned_but_never_eaten(client):
    """기존 4분면과의 결정적 차이 — 편성만 되고 취식 0인 메뉴가 보여야 한다.

    /menu-performance의 X축은 meal_log의 취식 발생 일수라 이런 메뉴는 아예
    나타나지 않는다. 그게 가장 강한 감편 신호인데 안 보이는 게 문제였다.
    """
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY, "제육볶음", "메인"),
                _plan_row(MONDAY + dt.timedelta(days=1), "아무도안먹은메뉴", "메인"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    _ingest_meal_log(client, "E1", "맛남", corner_name="한식", menu_name="제육볶음")

    data = _plan_performance(
        client,
        period_start=MONDAY.isoformat(),
        period_end=(MONDAY + dt.timedelta(days=5)).isoformat(),
    )
    by_name = {i["menu_name"]: i for i in data["items"]}
    assert "아무도안먹은메뉴" in by_name
    assert by_name["아무도안먹은메뉴"]["total_headcount"] == 0
    assert by_name["아무도안먹은메뉴"]["action"] == "취식 기록 없음"
    # 매칭 진단에도 잡혀야 한다 — 이름 불일치인지 담당자가 확인할 수 있게
    assert "아무도안먹은메뉴" in data["matching"]["plan_only"]
    assert data["matching"]["matched"] == 1


def test_plan_performance_excludes_side_dishes(client):
    """취식 데이터가 메인 기준이라 부찬을 넣으면 전부 취식 0이 되어 무의미하다."""
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY, "제육볶음", "메인"),
                _plan_row(MONDAY, "계란후라이", "부찬"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    data = _plan_performance(
        client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat()
    )
    assert [i["menu_name"] for i in data["items"]] == ["제육볶음"]


def test_plan_performance_counts_plan_appearances_not_intake_days(client):
    """편성 횟수는 식단표 기준 — 같은 날 여러 명이 먹어도 1회 편성이다."""
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY, "돈까스", "메인"),
                _plan_row(MONDAY + dt.timedelta(days=2), "돈까스", "메인"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    for i in range(4):
        _ingest_meal_log(client, f"E{i}", "맛남", corner_name="한식", menu_name="돈까스")

    data = _plan_performance(
        client,
        period_start=MONDAY.isoformat(),
        period_end=(MONDAY + dt.timedelta(days=5)).isoformat(),
    )
    row = next(i for i in data["items"] if i["menu_name"] == "돈까스")
    assert row["plan_count"] == 2  # 편성 2회
    assert row["total_headcount"] == 4  # 취식 4건
    assert row["headcount_per_plan"] == 2.0


def test_plan_performance_response_exposes_median_headcount_per_plan(client):
    """§80: X축이 '편성 횟수'에서 '1회 편성당 식수'로 바뀌면서 중앙값 필드명도
    median_plan_count → median_headcount_per_plan으로 바뀌었다."""
    client.post(
        "/api/ingest/weekly-menu",
        json={"rows": [_plan_row(MONDAY, "제육볶음", "메인")]},
        headers=AUTH_HEADERS,
    )
    _ingest_meal_log(client, "E1", "맛남", corner_name="한식", menu_name="제육볶음")

    data = _plan_performance(
        client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat()
    )
    assert "median_headcount_per_plan" in data
    assert "median_plan_count" not in data


def test_plan_performance_reports_log_only_menus(client):
    """취식은 있는데 그 기간 식단표에 MAIN으로 없는 메뉴도 알려준다."""
    client.post(
        "/api/ingest/weekly-menu",
        json={"rows": [_plan_row(MONDAY, "제육볶음", "메인")]},
        headers=AUTH_HEADERS,
    )
    _ingest_meal_log(client, "E1", "맛남", corner_name="한식", menu_name="식단표에없는메뉴")

    data = _plan_performance(
        client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat()
    )
    assert "식단표에없는메뉴" in data["matching"]["log_only"]


def test_request_scoped_caches_do_not_leak_across_requests(client):
    """성능용 세션 캐시(db.info)가 요청 경계를 넘어 오래된 값을 주면 안 된다.

    `_corner_id_by_menu_from_meal_log` 등은 180일 GROUP BY를 요청 단위로 캐시한다
    (2026-08 성능 개선). 캐시 수명이 요청보다 길어지면 새로 적재한 데이터가 화면에
    안 보이는 조용한 버그가 된다 — get_db가 요청마다 세션을 새로 만들기 때문에
    안전한데, 그 전제가 깨지면 이 테스트가 깨지도록 못박는다.
    """
    params = {
        "period_start": MONDAY.isoformat(),
        "period_end": (MONDAY + dt.timedelta(days=5)).isoformat(),
        "granularity": "daily",
        "group_by": "total",
    }
    before = client.get("/api/analysis/headcount-trend", params=params).json()
    total_before = sum(r["headcount"] for r in before)

    _ingest_meal_log(client, "CACHE1", "맛남", corner_name="한식", menu_name="제육볶음")

    after = client.get("/api/analysis/headcount-trend", params=params).json()
    total_after = sum(r["headcount"] for r in after)
    assert total_after == total_before + 1, "직전 요청의 캐시가 남아 새 취식이 안 보인다"


def test_reingest_without_replace_no_longer_duplicates_rows(client):
    """replace_existing을 안 켜도 같은 파일을 다시 올리면 행이 안 쌓인다.

    ⚠️ 이 테스트는 예전에 정반대를 주장했다 — "dedup이 없으므로 행이 쌓인다"를
    **의도된 성질로** 고정하고 있었다. 그런데 그 성질이 곧 2026-08 신고
    ("부찬이 두번씩 들어갔고")의 다른 얼굴이었다. 이제 이미 있는 행은 건너뛴다.

    replace_existing은 여전히 의미가 있다 — 그건 **없어진 메뉴를 지우는** 쪽이고,
    이 경로는 **있는 걸 또 넣지 않는** 쪽이다.
    """
    body = {"rows": [_plan_row(MONDAY, "제육볶음", "메인"), _plan_row(MONDAY, "계란후라이", "부찬")]}
    client.post("/api/ingest/weekly-menu", json=body, headers=AUTH_HEADERS)
    second = client.post("/api/ingest/weekly-menu", json=body, headers=AUTH_HEADERS)
    assert second.status_code == 200, second.text
    assert second.json()["skipped_duplicate"] == 2

    listed = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    ).json()
    slot = listed[0]
    assert slot["main"]["menu_name"] == "제육볶음"
    assert [s["menu_name"] for s in slot["sides"]] == ["계란후라이"]


def test_replace_existing_still_removes_menus_dropped_from_the_sheet(client):
    """중복 방지와 별개로, 식단표에서 빠진 메뉴는 교체 시 사라져야 한다."""
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [_plan_row(MONDAY, "제육볶음", "메인"), _plan_row(MONDAY, "계란후라이", "부찬")],
            "replace_existing": True,
        },
        headers=AUTH_HEADERS,
    )
    client.post(
        "/api/ingest/weekly-menu",
        json={"rows": [_plan_row(MONDAY, "제육볶음", "메인")], "replace_existing": True},
        headers=AUTH_HEADERS,
    )

    listed = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    ).json()
    assert listed[0]["sides"] == [], "식단표에서 빠진 부찬이 남아 있다"


def test_reingest_with_replace_existing_is_idempotent(client):
    """같은 payload를 몇 번 보내도 슬롯 내용이 그대로여야 한다."""
    body = {
        "rows": [_plan_row(MONDAY, "제육볶음", "메인"), _plan_row(MONDAY, "계란후라이", "부찬")],
        "replace_existing": True,
    }
    for _ in range(3):
        resp = client.post("/api/ingest/weekly-menu", json=body, headers=AUTH_HEADERS)
        assert resp.status_code == 200, resp.text

    listed = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    ).json()
    slot = listed[0]
    assert slot["main"]["menu_name"] == "제육볶음"
    assert [s["menu_name"] for s in slot["sides"]] == ["계란후라이"]


def test_replace_existing_preserves_manually_edited_rows(client, db_session):
    """관리자가 손으로 넣은 건강가든이 재업로드로 조용히 날아가면 안 된다."""
    _ingest_weekly_menu(client)
    corner_id = _corner_id(db_session, "한식")
    client.put(
        "/api/analysis/weekly-menu/health-garden",
        json={
            "plan_date": MONDAY.isoformat(),
            "corner_id": corner_id,
            "meal_type": "중식",
            "menu_names_raw": "구운채소",
        },
    )

    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [_plan_row(MONDAY, "돈까스", "메인"), _plan_row(MONDAY, "단무지", "부찬")],
            "replace_existing": True,
        },
        headers=AUTH_HEADERS,
    )

    listed = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    ).json()
    slot = next(s for s in listed if s["corner_id"] == corner_id)
    assert slot["main"]["menu_name"] == "돈까스"  # 식단표는 교체됨
    assert [i["menu_name"] for i in slot["health_garden"]] == ["구운채소"]  # 수기 입력은 보존


def test_origin_annotation_rows_no_longer_become_side_dishes(client):
    """파서 수정 전에는 `(계육-국산)`이 부찬으로 들어왔다.

    백엔드도 같은 정규화를 하므로, 설령 그런 이름이 적재 요청으로 들어와도
    메뉴명에서 원산지가 떨어져 유령 부찬이 생기지 않는지 확인한다.
    """
    client.post(
        "/api/ingest/weekly-menu",
        json={
            "rows": [
                _plan_row(MONDAY, "우삼겹구이(우육:호주산)", "메인"),
                _plan_row(MONDAY, "오징어(중국산)", "부찬"),
            ]
        },
        headers=AUTH_HEADERS,
    )
    listed = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    ).json()
    slot = listed[0]
    assert slot["main"]["menu_name"] == "우삼겹구이"
    assert [s["menu_name"] for s in slot["sides"]] == ["오징어"]


def test_rotation_reports_frequency_over_three_months_for_main_menu(client):
    """담당자 기준 "3개월에 2회까지 무난" — 메인이 3회면 과다로 표시된다.

    간격 기준(14일)만으로는 "14일은 넘겼지만 분기에 3번"이 안 잡힌다.
    """
    rows = [
        _plan_row(MONDAY - dt.timedelta(days=60), "갈비탕", "메인"),
        _plan_row(MONDAY - dt.timedelta(days=30), "갈비탕", "메인"),
        _plan_row(MONDAY, "갈비탕", "메인"),
    ]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    data = _rotation(client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat())
    item = next(i for i in data["items"] if i["menu_name"] == "갈비탕")
    assert item["window_count"] == 3
    assert item["window_max"] == 2
    assert item["over_frequency"] is True
    # 간격은 30일이라 "재편성 과다"(14일 기준)에는 안 걸린다 — 두 축이 다르다
    assert item["flag"] != "재편성 과다"


def test_rotation_frequency_threshold_is_looser_for_side_dishes(client):
    """김치·나물 같은 부찬은 자주 돌려쓰는 게 정상이다."""
    rows = [
        _plan_row(MONDAY - dt.timedelta(days=60), "김치", "부찬"),
        _plan_row(MONDAY - dt.timedelta(days=30), "김치", "부찬"),
        _plan_row(MONDAY, "김치", "부찬"),
    ]
    client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    data = _rotation(client, period_start=MONDAY.isoformat(), period_end=MONDAY.isoformat())
    item = next(i for i in data["items"] if i["menu_name"] == "김치")
    assert item["window_count"] == 3
    assert item["window_max"] == 6
    assert item["over_frequency"] is False


# ---------------------------------------------------------------------------
# 재적재 중복 사고 (2026-08 실사용 신고: "부찬이 두번씩 들어갔고")
# ---------------------------------------------------------------------------
# 원인: 교체 시 role_source=MANUAL 행을 안 지우는 건 맞는데, payload는 통째로
# 다시 넣어서 관리자가 손댄 메뉴가 슬롯에 두 벌씩 생겼다. 아래 테스트들은
# 수정 전에는 전부 깨진다.


def _slot(client, corner_id=None):
    listed = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()},
    ).json()
    if corner_id is None:
        return listed[0]
    return next(s for s in listed if s["corner_id"] == corner_id)


def test_reupload_does_not_duplicate_manually_reclassified_rows(client):
    """신고 재현 — 관리자가 주찬/부찬을 고친 뒤 재업로드하면 부찬이 두 배가 됐다.

    set_menu_role은 메인을 지정할 때 같은 슬롯의 **다른 MAIN들을 SIDE로 내리면서
    전부 MANUAL로 찍는다**(weekly_menu_review.py:151-153). 그래서 메인 하나만
    고쳐도 부찬 여러 개가 이 경로를 탄다.
    """
    body = {
        "rows": [
            _plan_row(MONDAY, "제육볶음", "메인"),
            _plan_row(MONDAY, "닭갈비", "메인"),  # 셀 병합 오판으로 메인이 둘
            _plan_row(MONDAY, "김치", "부찬"),
        ],
        "replace_existing": True,
    }
    client.post("/api/ingest/weekly-menu", json=body, headers=AUTH_HEADERS)

    # 관리자가 "닭갈비가 진짜 메인"이라고 고친다 → 제육볶음이 부찬(MANUAL)로 내려감
    slot = _slot(client)
    dakgalbi = next(
        i for i in [slot["main"], *slot["sides"]] if i and i["menu_name"] == "닭갈비"
    )
    resp = client.put(
        f"/api/analysis/weekly-menu/{dakgalbi['plan_id']}/role", json={"menu_role": "메인"}
    )
    assert resp.status_code == 200, resp.text

    before = _slot(client)
    before_names = sorted(i["menu_name"] for i in [before["main"], *before["sides"]] if i)

    # 같은 식단표를 다시 올린다
    client.post("/api/ingest/weekly-menu", json=body, headers=AUTH_HEADERS)

    after = _slot(client)
    after_names = sorted(i["menu_name"] for i in [after["main"], *after["sides"]] if i)
    assert after_names == before_names, f"재업로드로 행이 늘었다: {before_names} → {after_names}"
    assert len(after_names) == len(set(after_names)), f"같은 메뉴가 두 번 있다: {after_names}"


def test_reupload_keeps_the_manual_role_decision(client):
    """중복만 없애고 관리자 판단까지 되돌리면 안 된다 — 보존은 계속돼야 한다."""
    body = {
        "rows": [_plan_row(MONDAY, "제육볶음", "메인"), _plan_row(MONDAY, "닭갈비", "메인")],
        "replace_existing": True,
    }
    client.post("/api/ingest/weekly-menu", json=body, headers=AUTH_HEADERS)
    slot = _slot(client)
    dakgalbi = next(i for i in [slot["main"], *slot["sides"]] if i and i["menu_name"] == "닭갈비")
    client.put(f"/api/analysis/weekly-menu/{dakgalbi['plan_id']}/role", json={"menu_role": "메인"})

    client.post("/api/ingest/weekly-menu", json=body, headers=AUTH_HEADERS)

    assert _slot(client)["main"]["menu_name"] == "닭갈비", "관리자가 고른 메인이 되돌아갔다"


def test_reupload_reports_how_many_rows_it_skipped(client):
    """"재업로드했는데 왜 그대로지?"의 답이 응답에 있어야 한다."""
    body = {
        "rows": [_plan_row(MONDAY, "제육볶음", "메인"), _plan_row(MONDAY, "닭갈비", "메인")],
        "replace_existing": True,
    }
    client.post("/api/ingest/weekly-menu", json=body, headers=AUTH_HEADERS)
    slot = _slot(client)
    dakgalbi = next(i for i in [slot["main"], *slot["sides"]] if i and i["menu_name"] == "닭갈비")
    client.put(f"/api/analysis/weekly-menu/{dakgalbi['plan_id']}/role", json={"menu_role": "메인"})

    result = client.post("/api/ingest/weekly-menu", json=body, headers=AUTH_HEADERS).json()
    assert result["skipped_manual"] >= 1


def test_duplicate_rows_inside_one_payload_are_collapsed(client):
    """같은 부찬이 원본 셀에 두 번 적혀 있어도 한 행만 들어가야 한다.

    유니크 인덱스가 걸려 있으므로 여기서 안 거르면 정상 입력이 500으로 죽는다.
    """
    body = {
        "rows": [
            _plan_row(MONDAY, "제육볶음", "메인"),
            _plan_row(MONDAY, "김치", "부찬"),
            _plan_row(MONDAY, "김치", "부찬"),
        ],
        "replace_existing": True,
    }
    resp = client.post("/api/ingest/weekly-menu", json=body, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text
    assert resp.json()["skipped_duplicate"] == 1
    assert [s["menu_name"] for s in _slot(client)["sides"]] == ["김치"]


def test_duplicate_payload_rows_are_collapsed_even_without_replace(client):
    """replace_existing를 안 켜도 payload 내 중복은 제약에 걸리면 안 된다."""
    body = {
        "rows": [_plan_row(MONDAY, "제육볶음", "메인"), _plan_row(MONDAY, "김치", "부찬")] * 2
    }
    resp = client.post("/api/ingest/weekly-menu", json=body, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text
    assert resp.json()["skipped_duplicate"] == 2


# ---------------------------------------------------------------------------
# 표기가 달라 같은 메뉴가 갈라지던 문제 (2026-08 "연어파피요트" 신고)
# ---------------------------------------------------------------------------


def test_menu_name_spacing_difference_still_matches_the_same_menu(client, db_session):
    """식단표는 "연어 파피요트", POS는 "연어파피요트"로 와도 같은 메뉴여야 한다.

    예전엔 menu_name 정확 일치로 찾아서 별개 행이 됐고, 매칭 진단에서 같은 이름이
    plan_only와 log_only에 동시에 떴다.
    """
    client.post(
        "/api/ingest/weekly-menu",
        json={"rows": [_plan_row(MONDAY, "연어 파피요트", "메인")]},
        headers=AUTH_HEADERS,
    )
    _ingest_meal_log(client, "E12345", "맛남", menu_name="연어파피요트")

    from app.models.master import MenuMaster

    rows = [m for m in db_session.query(MenuMaster).all() if "파피요트" in m.menu_name]
    assert len(rows) == 1, f"같은 메뉴가 여러 행으로 갈라졌다: {[m.menu_name for m in rows]}"


def test_origin_annotation_longer_than_six_chars_is_still_stripped(client, db_session):
    """`노르웨이자연산`(7자)을 못 떼서 메뉴가 갈라졌던 경로."""
    client.post(
        "/api/ingest/weekly-menu",
        json={"rows": [_plan_row(MONDAY, "연어파피요트(연어:노르웨이자연산)", "메인")]},
        headers=AUTH_HEADERS,
    )
    from app.models.master import MenuMaster

    rows = [m for m in db_session.query(MenuMaster).all() if "파피요트" in m.menu_name]
    assert [m.menu_name for m in rows] == ["연어파피요트"]


def test_health_garden_menu_can_later_appear_in_the_weekly_sheet(client, db_session):
    """건강가든으로 만든 메뉴가 나중에 식단표로 들어와도 터지면 안 된다.

    예전엔 건강가든 경로가 get_or_create_menu를 안 써서 match_key가 NULL로 남았고,
    나중에 같은 이름이 들어오면 조회가 못 찾아 menu_name unique 위반이 났다
    (2026-08, 병합 스크립트 오류와 같은 뿌리).
    """
    _ingest_weekly_menu(client)
    corner_id = _corner_id(db_session, "한식")
    resp = client.put(
        "/api/analysis/weekly-menu/health-garden",
        json={
            "plan_date": MONDAY.isoformat(),
            "corner_id": corner_id,
            "meal_type": "중식",
            "menu_names_raw": "구운채소",
        },
    )
    assert resp.status_code == 200, resp.text

    # 같은 이름이 식단표로 들어온다 — 여기서 예전엔 IntegrityError가 났다.
    resp = client.post(
        "/api/ingest/weekly-menu",
        json={"rows": [_plan_row(MONDAY + dt.timedelta(days=1), "구운채소", "부찬")]},
        headers=AUTH_HEADERS,
    )
    assert resp.status_code == 200, resp.text

    from app.models.master import MenuMaster

    rows = [m for m in db_session.query(MenuMaster).all() if m.menu_name == "구운채소"]
    assert len(rows) == 1
    assert rows[0].match_key == "구운채소", "match_key가 안 채워졌다"


def test_ingest_weather_csv_upserts_by_date(client, db_session):
    from app.models.stats import DailyWeather

    rows = [
        {"stat_date": "2026-08-01", "precip_mm": 12.5, "avg_temp_c": 24.3},
        {"stat_date": "2026-08-02", "precip_mm": None, "avg_temp_c": 26.0},
    ]
    resp = client.post("/api/ingest/weather-csv", json={"rows": rows}, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text
    assert resp.json() == {"received": 2, "upserted": 2}

    stored = {w.stat_date: w for w in db_session.query(DailyWeather).all()}
    assert stored[dt.date(2026, 8, 1)].had_rain is True
    assert stored[dt.date(2026, 8, 2)].had_rain is False
    assert stored[dt.date(2026, 8, 1)].source == "csv_import"

    # 같은 날짜를 다시 올리면 갱신되고 중복 행이 쌓이지 않는다.
    resp = client.post(
        "/api/ingest/weather-csv",
        json={"rows": [{"stat_date": "2026-08-01", "precip_mm": 0, "avg_temp_c": 30.0}]},
        headers=AUTH_HEADERS,
    )
    assert resp.status_code == 200
    assert db_session.query(DailyWeather).count() == 2
    updated = db_session.get(DailyWeather, dt.date(2026, 8, 1))
    db_session.refresh(updated)
    assert updated.had_rain is False
    assert updated.avg_temp_c == 30.0


def test_ingest_weather_csv_requires_token(client):
    resp = client.post("/api/ingest/weather-csv", json={"rows": []})
    assert resp.status_code == 401


def _seed_menu_rain_vs_normal(client, db_session, menu_name: str, start_date: dt.date, employee_prefix: str) -> None:
    """§71: 비오는 날 5일(6명씩) vs 평상시 5일(2명씩) 같은 메뉴로 채운다 —
    weather-event-ranking/predicted-impact weather_reference 테스트 공용 시딩."""
    from app.models.stats import DailyWeather

    rain_days = [start_date + dt.timedelta(days=i) for i in range(5)]
    normal_days = [start_date + dt.timedelta(days=i) for i in range(5, 10)]
    emp_n = 0
    for d in rain_days:
        db_session.add(
            DailyWeather(
                stat_date=d, had_rain=True, precip_mm=8.0, snow_cm=0.0, max_temp_c=24.0, min_temp_c=18.0
            )
        )
        for _ in range(6):
            _ingest_meal_log(client, f"{employee_prefix}{emp_n}", "맛남", eaten_date=d, menu_name=menu_name)
            emp_n += 1
    for d in normal_days:
        db_session.add(
            DailyWeather(
                stat_date=d, had_rain=False, precip_mm=0.0, snow_cm=0.0, max_temp_c=24.0, min_temp_c=18.0
            )
        )
        for _ in range(2):
            _ingest_meal_log(client, f"{employee_prefix}{emp_n}", "맛남", eaten_date=d, menu_name=menu_name)
            emp_n += 1
    db_session.commit()


def test_menu_weather_event_ranking_surfaces_high_diff_menu(client, db_session):
    """비 오는 날 유독 잘 나가는 메뉴가 랭킹에 뜨고 diff 부호/값이 맞는지 확인
    (담당자 요청 예시 그대로: "비오면 김치찌개가 평소보다 많이 찾았다")."""
    _seed_menu_rain_vs_normal(client, db_session, "김치찌개", MONDAY, "E")

    resp = client.get(
        "/api/analysis/menu-performance/weather-event-ranking",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": (MONDAY + dt.timedelta(days=9)).isoformat(),
            "event": "비",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["event"] == "비"
    assert body["actual_metric_label"] == "평균 강수량(mm)"
    row = next(r for r in body["rows"] if r["menu_name"] == "김치찌개")
    assert row["low_sample"] is False
    assert row["event_days"] == 5
    assert row["event_avg_headcount"] == 6.0
    assert row["diff_vs_normal"] == 4.0  # 비 오는 날 6명 - 평상시 2명
    assert row["actual_avg"] == 8.0  # _seed_menu_rain_vs_normal의 비 오는 날 precip_mm
    assert body["rows"][0]["menu_name"] == "김치찌개"  # |diff| 내림차순 1위


def test_menu_weather_event_ranking_flags_low_sample(client, db_session):
    from app.models.stats import DailyWeather

    rain_day = MONDAY
    db_session.add(
        DailyWeather(
            stat_date=rain_day, had_rain=True, precip_mm=5.0, snow_cm=0.0, max_temp_c=24.0, min_temp_c=18.0
        )
    )
    db_session.commit()
    _ingest_meal_log(client, "E1", "맛남", eaten_date=rain_day, menu_name="우동")

    resp = client.get(
        "/api/analysis/menu-performance/weather-event-ranking",
        params={"period_start": rain_day.isoformat(), "period_end": rain_day.isoformat(), "event": "비"},
    )
    assert resp.status_code == 200, resp.text
    row = next(r for r in resp.json()["rows"] if r["menu_name"] == "우동")
    assert row["low_sample"] is True
    assert row["diff_vs_normal"] is None
    # 표본이 부족해 diff는 못 내도 실측 강수량 자체는 그대로 보여준다(low_sample과 무관).
    assert row["actual_avg"] == 5.0


def test_menu_weather_event_ranking_rejects_invalid_event(client):
    resp = client.get(
        "/api/analysis/menu-performance/weather-event-ranking",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "event": "이상한값"},
    )
    assert resp.status_code == 400


def test_menu_weather_event_ranking_excludes_micam_hall_corner(client, db_session):
    """§76: 담당자 요청("미캠회관 코너도 제외해줘") — 미캠회관(전골) 코너에서
    나간 메뉴는 날씨유형 랭킹에서 빠지고, 다른 코너 메뉴는 그대로 남아야 한다."""
    # 대조군이 이미 비 오는 날 5일치 DailyWeather를 시딩하므로, 미캠회관(전골)
    # 메뉴도 같은 날짜에 얹어 중복 DailyWeather insert를 피한다.
    _seed_menu_rain_vs_normal(client, db_session, "김치찌개", MONDAY, "K")

    rain_days = [MONDAY + dt.timedelta(days=i) for i in range(5)]
    emp_n = 0
    for d in rain_days:
        for _ in range(6):
            _ingest_meal_log(
                client, f"H{emp_n}", "맛남", eaten_date=d, menu_name="전골", corner_name="미캠회관(전골)"
            )
            emp_n += 1

    resp = client.get(
        "/api/analysis/menu-performance/weather-event-ranking",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": (MONDAY + dt.timedelta(days=9)).isoformat(),
            "event": "비",
        },
    )
    assert resp.status_code == 200, resp.text
    rows = resp.json()["rows"]
    assert all(r["menu_name"] != "전골" for r in rows)
    assert any(r["menu_name"] == "김치찌개" for r in rows)


def test_menu_weather_correlation_ranking_surfaces_positive_correlation(client, db_session):
    """§81: 기온이 오를수록 식수가 느는 메뉴를 시딩하면 양의 상관계수로 뜨는지
    확인 — weather-event-ranking과 달리 연속값이라 6일 모두 다른 기온/식수를 준다."""
    from app.models.stats import DailyWeather

    days = [MONDAY + dt.timedelta(days=i) for i in range(6)]
    temps = [10.0, 15.0, 20.0, 25.0, 30.0, 35.0]
    headcounts = [1, 2, 3, 4, 5, 6]
    emp_n = 0
    for d, temp, headcount in zip(days, temps, headcounts):
        db_session.add(
            DailyWeather(stat_date=d, had_rain=False, precip_mm=0.0, snow_cm=0.0, max_temp_c=temp, min_temp_c=temp - 8)
        )
        for _ in range(headcount):
            _ingest_meal_log(client, f"T{emp_n}", "맛남", eaten_date=d, menu_name="냉면")
            emp_n += 1
    db_session.commit()

    resp = client.get(
        "/api/analysis/menu-performance/weather-correlation-ranking",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": (MONDAY + dt.timedelta(days=5)).isoformat(),
            "metric": "max_temp_c",
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["metric"] == "max_temp_c"
    assert body["metric_label"] == "최고기온(℃)"
    row = next(r for r in body["rows"] if r["menu_name"] == "냉면")
    assert row["sample_size"] == 6
    assert row["correlation"] == 1.0
    assert body["rows"][0]["menu_name"] == "냉면"  # 상관계수 내림차순 1위


def test_menu_weather_correlation_ranking_excludes_low_sample_menu(client, db_session):
    """§81: min_days 미만인 메뉴는 우연한 상관관계로 보고 응답에서 빠진다."""
    from app.models.stats import DailyWeather

    days = [MONDAY + dt.timedelta(days=i) for i in range(2)]
    for i, d in enumerate(days):
        db_session.add(DailyWeather(stat_date=d, had_rain=False, precip_mm=0.0, max_temp_c=20.0 + i, min_temp_c=10.0))
        _ingest_meal_log(client, f"L{i}", "맛남", eaten_date=d, menu_name="비빔밥")
    db_session.commit()

    resp = client.get(
        "/api/analysis/menu-performance/weather-correlation-ranking",
        params={
            "period_start": MONDAY.isoformat(),
            "period_end": (MONDAY + dt.timedelta(days=1)).isoformat(),
            "metric": "max_temp_c",
            "min_days": 5,
        },
    )
    assert resp.status_code == 200, resp.text
    assert all(r["menu_name"] != "비빔밥" for r in resp.json()["rows"])


def test_predicted_impact_includes_weather_reference_when_history_exists(client, db_session):
    plan_date = MONDAY + dt.timedelta(days=30)
    _seed_menu_rain_vs_normal(client, db_session, "김치찌개", MONDAY, "W")

    rows = [
        {
            "plan_date": plan_date.isoformat(),
            "meal_type": "중식",
            "corner_name": "한식",
            "menu_name": "김치찌개",
            "menu_role": "메인",
            "source_row_raw": "김치찌개",
        }
    ]
    resp = client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text

    resp = client.get(
        "/api/analysis/weekly-menu",
        params={"period_start": plan_date.isoformat(), "period_end": plan_date.isoformat()},
    )
    main_plan_id = next(s for s in resp.json() if s["corner_name"] == "한식")["main"]["plan_id"]

    resp = client.get(f"/api/analysis/weekly-menu/{main_plan_id}/predicted-impact")
    assert resp.status_code == 200, resp.text
    weather_reference = {r["event"]: r for r in resp.json()["weather_reference"]}
    assert weather_reference["비"]["diff_vs_normal"] == 4.0
    assert weather_reference["비"]["low_sample"] is False


def test_predicted_impact_weather_reference_empty_without_history(client):
    """이 슬롯 이전 이력이 전혀 없으면(날씨든 식수든) 에러 없이 빈 리스트로
    조용히 빠진다 — 미설정 시 기능 비활성화 관례(weather_client.py와 동일)."""
    _ingest_weekly_menu(client)  # 제육볶음(메인), 한식, MONDAY — 이전 이력 없음
    for i in range(3):
        _ingest_meal_log(client, f"P{i}", "맛남", menu_name="제육볶음")

    resp = client.get(
        "/api/analysis/weekly-menu", params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat()}
    )
    main_plan_id = next(s for s in resp.json() if s["corner_name"] == "한식")["main"]["plan_id"]

    resp = client.get(f"/api/analysis/weekly-menu/{main_plan_id}/predicted-impact")
    assert resp.status_code == 200, resp.text
    assert resp.json()["weather_reference"] == []


def test_weather_event_ranking_flags_extended_fields_missing(client, db_session):
    """§72: snow_cm/max_temp_c/min_temp_c가 전부 NULL인(§71 배포 전에 이미
    백필된) daily_weather만 있으면, "그런 날이 없어서"와 구분해 화면이 재백필
    필요를 알 수 있도록 extended_fields_missing이 True여야 한다."""
    from app.models.stats import DailyWeather

    db_session.add(
        DailyWeather(stat_date=MONDAY, had_rain=True, precip_mm=8.0, snow_cm=None, max_temp_c=None, min_temp_c=None)
    )
    db_session.commit()
    _ingest_meal_log(client, "E1", "맛남", eaten_date=MONDAY, menu_name="김치찌개")

    resp = client.get(
        "/api/analysis/menu-performance/weather-event-ranking",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "event": "폭설"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["rows"] == []
    assert body["extended_fields_missing"] is True


def test_weather_event_ranking_extended_fields_missing_false_once_backfilled(client, db_session):
    """확장 필드가 하나라도 채워진 날이 있으면(재백필 완료) False여야 한다."""
    from app.models.stats import DailyWeather

    db_session.add(
        DailyWeather(stat_date=MONDAY, had_rain=True, precip_mm=8.0, snow_cm=0.0, max_temp_c=24.0, min_temp_c=18.0)
    )
    db_session.commit()

    resp = client.get(
        "/api/analysis/menu-performance/weather-event-ranking",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "event": "폭설"},
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["extended_fields_missing"] is False


def _seed_menu_summer_vs_fall(client, menu_name: str, employee_prefix: str) -> None:
    """§72: 여름(7월) 5일(6명씩) vs 가을(10월) 5일(2명씩) 같은 메뉴로 채운다 —
    season-ranking 테스트 공용 시딩."""
    summer_days = [dt.date(2026, 7, 1) + dt.timedelta(days=i) for i in range(5)]
    fall_days = [dt.date(2026, 10, 1) + dt.timedelta(days=i) for i in range(5)]
    emp_n = 0
    for d in summer_days:
        for _ in range(6):
            _ingest_meal_log(client, f"{employee_prefix}{emp_n}", "맛남", eaten_date=d, menu_name=menu_name)
            emp_n += 1
    for d in fall_days:
        for _ in range(2):
            _ingest_meal_log(client, f"{employee_prefix}{emp_n}", "맛남", eaten_date=d, menu_name=menu_name)
            emp_n += 1


def test_menu_season_ranking_surfaces_summer_favorite(client):
    """담당자 요청 예시("냉면은 여름에") 그대로 — 여름에 유독 잘 나가는 메뉴가
    상위에, diff_vs_overall 부호/값이 맞는지 확인."""
    _seed_menu_summer_vs_fall(client, "냉면", "S")

    resp = client.get(
        "/api/analysis/menu-performance/season-ranking",
        params={"period_start": "2026-07-01", "period_end": "2026-10-05", "season": "여름"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["season"] == "여름"
    row = next(r for r in body["rows"] if r["menu_name"] == "냉면")
    assert row["low_sample"] is False
    assert row["season_days"] == 5
    assert row["season_avg_headcount"] == 6.0
    # 전체 평균 4.0(여름 6명×5일 + 가을 2명×5일) 대비 여름은 +2.0
    assert row["diff_vs_overall"] == 2.0
    assert body["rows"][0]["menu_name"] == "냉면"


def test_menu_season_ranking_excludes_micam_hall_corner(client):
    """§76: 미캠회관(전골) 코너의 메뉴는 계절 랭킹에서도 빠져야 한다 —
    날씨유형·계절 랭킹이 같은 집계 헬퍼(_headcount_by_date_by_menu_bulk)를
    공유하므로 한 곳만 고쳐도 둘 다 적용된다."""
    summer_days = [dt.date(2026, 7, 1) + dt.timedelta(days=i) for i in range(5)]
    emp_n = 0
    for d in summer_days:
        for _ in range(6):
            _ingest_meal_log(
                client, f"J{emp_n}", "맛남", eaten_date=d, menu_name="전골", corner_name="미캠회관(전골)"
            )
            emp_n += 1
    _seed_menu_summer_vs_fall(client, "냉면", "N")

    resp = client.get(
        "/api/analysis/menu-performance/season-ranking",
        params={"period_start": "2026-07-01", "period_end": "2026-10-05", "season": "여름"},
    )
    assert resp.status_code == 200, resp.text
    rows = resp.json()["rows"]
    assert all(r["menu_name"] != "전골" for r in rows)
    assert any(r["menu_name"] == "냉면" for r in rows)


def test_menu_season_ranking_flags_low_sample(client):
    _ingest_meal_log(client, "F1", "맛남", eaten_date=dt.date(2026, 10, 1), menu_name="어묵탕")

    resp = client.get(
        "/api/analysis/menu-performance/season-ranking",
        params={"period_start": "2026-10-01", "period_end": "2026-10-01", "season": "가을"},
    )
    assert resp.status_code == 200, resp.text
    row = next(r for r in resp.json()["rows"] if r["menu_name"] == "어묵탕")
    assert row["low_sample"] is True
    assert row["diff_vs_overall"] is None


def test_menu_season_ranking_rejects_invalid_season(client):
    resp = client.get(
        "/api/analysis/menu-performance/season-ranking",
        params={"period_start": MONDAY.isoformat(), "period_end": MONDAY.isoformat(), "season": "장마"},
    )
    assert resp.status_code == 400


# ---------------------------------------------------------------------------
# §77~§78: 주간 식단표 규칙 검증 — 담당자가 준 4개 기준(해장/면류/매운빨간국물/
# 최근 저조 식수 재편성 금지)이 API 레벨에서 정확히 판정되는지 확인한다.
# §78부터 해장/면류/매운빨간국물은 요일별(하루 기준, 주중만) 응답이라
# hangover/noodle/spicy_red_broth는 날짜별 결과 배열이다.
# ---------------------------------------------------------------------------


def _daily_result_for(results: list[dict], plan_date: dt.date) -> dict:
    return next(r for r in results if r["plan_date"] == plan_date.isoformat())


def test_weekly_menu_plan_rule_check_flags_noodle_overage_on_a_single_day(client):
    """5개를 월요일 하루에 몰아넣으면 그날이 위반돼야 한다 — 하루 기준 판정의
    핵심 케이스. §79에서 "짬뽕"이 해장 키워드에도 추가돼 hangover 어서션과
    겹치므로, 순수 면류만 걸리는 "쫄면"으로 바꿨다(면류 판정 자체는 동일)."""
    noodle_menus = ["라면", "우동", "짜장면", "쫄면", "냉면"]
    rows = [_plan_row(MONDAY, name, "메인") for name in noodle_menus]
    resp = client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text

    resp = client.get(
        "/api/analysis/weekly-menu/plan-rule-check",
        params={"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=4)).isoformat()},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    monday_hangover = _daily_result_for(body["hangover"], MONDAY)
    monday_noodle = _daily_result_for(body["noodle"], MONDAY)
    monday_spicy = _daily_result_for(body["spicy_red_broth"], MONDAY)
    assert monday_hangover["ok"] is False
    assert monday_hangover["count"] == 0
    assert monday_noodle["ok"] is False
    assert monday_noodle["count"] == 5
    assert monday_noodle["limit"] == 4
    assert len(monday_noodle["matches"]) == 5
    assert monday_spicy["ok"] is True


def test_weekly_menu_plan_rule_check_noodle_spread_across_days_all_pass(client):
    """같은 5개를 월~금 하루 1개씩 나눠 편성하면 매일 통과해야 한다 — §77(주
    전체 합산)에서 §78(하루 기준)로 바뀐 걸 직접 보여주는 회귀 방지 테스트."""
    noodle_menus = ["라면", "우동", "짜장면", "짬뽕", "냉면"]
    rows = [_plan_row(MONDAY + dt.timedelta(days=i), name, "메인") for i, name in enumerate(noodle_menus)]
    resp = client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text

    resp = client.get(
        "/api/analysis/weekly-menu/plan-rule-check",
        params={"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=4)).isoformat()},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["noodle"]) == 5
    assert all(r["ok"] is True and r["count"] == 1 for r in body["noodle"])


def test_weekly_menu_plan_rule_check_passes_with_hangover_and_low_noodle_count(client):
    rows = [_plan_row(MONDAY, "황태해장국", "메인")]
    resp = client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text

    resp = client.get(
        "/api/analysis/weekly-menu/plan-rule-check",
        params={"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=4)).isoformat()},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert _daily_result_for(body["hangover"], MONDAY)["ok"] is True
    assert _daily_result_for(body["noodle"], MONDAY)["ok"] is True
    assert _daily_result_for(body["spicy_red_broth"], MONDAY)["ok"] is True
    assert body["low_headcount_reuse"]["ok"] is True


def test_weekly_menu_plan_rule_check_excludes_saturday_from_daily_rules(client):
    """주중만 본다 — 토요일에 면류를 몰아넣어도 규칙 위반 목록에 안 잡혀야 한다."""
    saturday = MONDAY + dt.timedelta(days=5)
    noodle_menus = ["라면", "우동", "짜장면", "짬뽕", "냉면"]
    rows = [_plan_row(saturday, name, "메인") for name in noodle_menus]
    resp = client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text

    resp = client.get(
        "/api/analysis/weekly-menu/plan-rule-check",
        params={"period_start": MONDAY.isoformat(), "period_end": saturday.isoformat()},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert all(r["plan_date"] != saturday.isoformat() for r in body["noodle"])


def test_weekly_menu_plan_rule_check_flags_low_headcount_reuse_except_exempt_corners(client):
    """최근 저조 식수(200식 이하) 메뉴는 재편성 위반으로 뜨지만, 미캠회관(전골)
    같은 예외 코너 메뉴는 같은 조건이어도 위반 목록에서 빠져야 한다."""
    history_end = MONDAY - dt.timedelta(days=1)
    history_days = [history_end - dt.timedelta(days=i) for i in range(5)]

    for i, d in enumerate(history_days):
        _ingest_meal_log(client, f"L{i}", "맛남", eaten_date=d, menu_name="고기전골", corner_name="한식")
    for i, d in enumerate(history_days):
        _ingest_meal_log(
            client, f"M{i}", "맛남", eaten_date=d, menu_name="미캠전골", corner_name="미캠회관(전골)"
        )

    rows = [
        _plan_row(MONDAY, "고기전골", "메인", corner_name="한식"),
        _plan_row(MONDAY, "미캠전골", "메인", corner_name="미캠회관(전골)"),
    ]
    resp = client.post("/api/ingest/weekly-menu", json={"rows": rows}, headers=AUTH_HEADERS)
    assert resp.status_code == 200, resp.text

    resp = client.get(
        "/api/analysis/weekly-menu/plan-rule-check",
        params={"period_start": MONDAY.isoformat(), "period_end": (MONDAY + dt.timedelta(days=4)).isoformat()},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["low_headcount_reuse"]["ok"] is False
    names = {v["menu_name"] for v in body["low_headcount_reuse"]["violations"]}
    assert "고기전골" in names
    assert "미캠전골" not in names


def test_recompute_llm_analyses_populates_menu_highlight_cause(client):
    """§78: "현황에서 메뉴하이라이트llm에 분석이 아무것도 없어" — 로컬처럼
    새벽 배치 스케줄러가 안 떠 있어도 이 엔드포인트로 캐시를 수동으로 채우면
    메뉴 하이라이트 응답에 cause가 채워져야 한다. 테스트 환경은 LLM 미설정
    이라 폴백 문구("...미설정...")가 캐시된다."""
    today = dt.date.today()
    recent_monday = today - dt.timedelta(days=today.weekday())
    prior_monday = recent_monday - dt.timedelta(days=7)

    for i in range(2):
        _ingest_meal_log(client, f"R{i}", "맛남", eaten_date=recent_monday, menu_name="회복원산지찌개")
    for i in range(2):
        _ingest_meal_log(client, f"P{i}", "개선", eaten_date=prior_monday, menu_name="회복원산지찌개")

    resp = client.post(
        "/api/analysis/llm-analyses/recompute",
        params={"period_start": today.isoformat(), "period_end": today.isoformat()},
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["menu_trend"] >= 1

    resp = client.get("/api/dashboard/menu-highlights")
    assert resp.status_code == 200, resp.text
    entries = [*resp.json()["rising"], *resp.json()["falling"]]
    target = next(e for e in entries if e["menu_name"] == "회복원산지찌개")
    assert "cause" in target
    assert "미설정" in target["cause"]
